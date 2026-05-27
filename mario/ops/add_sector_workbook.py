"""Reader and writer helpers for the add-sectors workbook workflow."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils.cell import range_boundaries

from mario.log_exc.exceptions import WrongExcelFormat, WrongInput
from mario.model.conventions import IOT, SUT, _MASTER_INDEX
from mario.ops.add_sector_specs import (
    ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS,
    ADD_SECTOR_SPLIT_EXCLUSION_SHEET,
    ADD_SECTOR_SPLIT_OUTPUT_COLUMNS,
    ADD_SECTOR_SPLIT_OUTPUT_SHEET,
    ADD_SECTOR_SPLIT_TOLERANCE_COLUMNS,
    ADD_SECTOR_SPLIT_TOLERANCE_DEFAULTS,
    ADD_SECTOR_SPLIT_TOLERANCE_SHEET,
    ADD_SECTOR_SPLIT_TRADE_COLUMNS,
    ADD_SECTOR_SPLIT_TRADE_SHEET,
    ADVANCED_ADD_SECTOR_DB_UNITS_SHEET,
    ADVANCED_ADD_SECTOR_FACTORS_CLUSTERS_SHEET,
    ADVANCED_ADD_SECTOR_INVENTORY_SHEET_COLUMNS,
    ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_COLUMNS,
    ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET,
    ADVANCED_ADD_SECTOR_MASTER_SHEET,
    ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS,
    ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_COLUMNS,
    ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET,
    ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET,
    ADVANCED_ADD_SECTOR_UNCERTAINTY_COLUMNS,
    ADVANCED_ADD_SECTOR_UNCERTAINTY_PARAMETERS,
)


@dataclass(frozen=True)
class AddSectorWorkbook:
    """Normalized view of one add-sectors workbook.

    The workbook stores:
    - one master sheet describing the items to add and their workbook-level options;
    - one regions-cluster sheet;
    - one item-cluster sheet;
    - an optional uncertainty-value sheet;
    - one or more inventory sheets referenced by the master sheet.

    This dataclass keeps those pieces together after parsing so ``Database`` can
    attach them to the live object and the engine can consume them without
    repeatedly reopening the Excel file.
    """

    table: str
    master_sheet: pd.DataFrame
    regions_clusters: dict[str, list[str]]
    item_clusters: dict[str, list[str]]
    factors_clusters: dict[str, list[str]]
    uncertainty_values: dict[str, float]
    inventories_by_sheet: dict[str, pd.DataFrame]
    split_info: dict[str, pd.DataFrame] | None = None


def build_add_sector_master_sheet(
    table: str,
    new_items: list[str],
    regions: list[str],
    *,
    item: str | None = None,
) -> pd.DataFrame:
    """Build a prefilled master sheet for a new add-sectors workbook.

    Parameters
    ----------
    table:
        ``"IOT"`` or ``"SUT"``.
    new_items:
        Item names to pre-populate in the master sheet. An empty list is valid
        and produces an empty master sheet with only the expected columns.
    regions:
        Regions to pre-populate alongside ``new_items``.
    item:
        For SUT workbooks, controls whether the same ``new_items`` should be
        written as activities, commodities, or both.
    """

    columns = list(ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[table].values())
    rows: list[dict[str, Any]] = []

    counter = 1
    if table == IOT:
        for region in regions:
            for sector in new_items:
                row = {column: "" for column in columns}
                row[_MASTER_INDEX["r"]] = region
                row[_MASTER_INDEX["s"]] = sector
                row["Inventory sheet"] = f"INV_{counter:03d}"
                row["Add or Split"] = "Add"
                rows.append(row)
                counter += 1
    else:
        for region in regions:
            for value in new_items:
                row = {column: "" for column in columns}
                row[_MASTER_INDEX["r"]] = region
                if item is None:
                    row[_MASTER_INDEX["a"]] = value
                    row[_MASTER_INDEX["c"]] = value
                elif item == _MASTER_INDEX["a"]:
                    row[_MASTER_INDEX["a"]] = value
                    row[_MASTER_INDEX["c"]] = ""
                elif item == _MASTER_INDEX["c"]:
                    row[_MASTER_INDEX["a"]] = ""
                    row[_MASTER_INDEX["c"]] = value
                else:
                    raise WrongInput(
                        f"For SUT add-sectors workbooks, item should be "
                        f"{_MASTER_INDEX['a']}, {_MASTER_INDEX['c']} or None."
                    )
                row["Inventory sheet"] = f"INV_{counter:03d}"
                rows.append(row)
                counter += 1

    return pd.DataFrame(rows, columns=columns)


def build_regions_clusters_sheet(instance) -> pd.DataFrame:
    """Build the default regions-cluster sheet."""

    return pd.DataFrame(
        instance.get_index(_MASTER_INDEX["r"]),
        columns=ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_COLUMNS,
    )


def build_items_clusters_sheet() -> pd.DataFrame:
    """Build the default item-cluster sheet."""

    return pd.DataFrame(columns=ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_COLUMNS)


def build_uncertainties_sheet() -> pd.DataFrame:
    """Build the default uncertainty-values sheet."""

    return pd.DataFrame(
        {
            ADVANCED_ADD_SECTOR_UNCERTAINTY_COLUMNS[0]: list(
                ADVANCED_ADD_SECTOR_UNCERTAINTY_PARAMETERS.keys()
            ),
            ADVANCED_ADD_SECTOR_UNCERTAINTY_COLUMNS[1]: list(
                ADVANCED_ADD_SECTOR_UNCERTAINTY_PARAMETERS.values()
            ),
        }
    )


def build_inventory_template() -> pd.DataFrame:
    """Build one empty inventory sheet."""

    return pd.DataFrame(
        columns=list(ADVANCED_ADD_SECTOR_INVENTORY_SHEET_COLUMNS.values())
    )


def build_split_total_outputs_template() -> pd.DataFrame:
    """Build the split-specific total-output sheet."""

    return pd.DataFrame(columns=list(ADD_SECTOR_SPLIT_OUTPUT_COLUMNS.values()))


def build_split_trades_template() -> pd.DataFrame:
    """Build the split-specific trade sheet."""

    return pd.DataFrame(columns=list(ADD_SECTOR_SPLIT_TRADE_COLUMNS.values()))


def build_split_exclusions_template() -> pd.DataFrame:
    """Build the split-specific exclusions sheet."""

    return pd.DataFrame(columns=list(ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS.values()))


def build_split_tolerances_template() -> pd.DataFrame:
    """Build the split-specific tolerance sheet with default scalar rows."""

    return pd.DataFrame(
        ADD_SECTOR_SPLIT_TOLERANCE_DEFAULTS,
        columns=list(ADD_SECTOR_SPLIT_TOLERANCE_COLUMNS.values()),
    )


def build_db_units_sheet(instance) -> pd.DataFrame:
    """Build a flat view of database units for workbook reference."""

    rows = []
    if instance.table_type == SUT:
        keys = [
            _MASTER_INDEX["a"],
            _MASTER_INDEX["c"],
            _MASTER_INDEX["f"],
            _MASTER_INDEX["k"],
        ]
    else:
        keys = [_MASTER_INDEX["s"], _MASTER_INDEX["f"], _MASTER_INDEX["k"]]

    for key in keys:
        units = instance.units[key]
        for item_name, row in units.iterrows():
            rows.append(
                {
                    "Item type": key,
                    "DB Item": item_name,
                    "Unit": row["unit"],
                }
            )

    return pd.DataFrame(rows, columns=["Item type", "DB Item", "Unit"])


def write_add_sector_workbook(
    instance,
    path: str | Path,
    *,
    new_items: list[str],
    regions: list[str],
    item: str | None = None,
    redefine_uncertainties: bool = False,
) -> None:
    """Write an add-sectors workbook to disk.

    The workbook always includes the structural sheets needed by the current
    ``add_sectors`` workflow. When ``new_items`` and ``regions`` are provided,
    the master sheet and the inventory sheets are pre-populated accordingly.
    When they are empty, the file acts as a blank template that users can fill
    manually.
    """

    table = instance.table_type
    master = build_add_sector_master_sheet(table, new_items, regions, item=item)
    regions_clusters = build_regions_clusters_sheet(instance)
    items_clusters = build_items_clusters_sheet()
    db_units = build_db_units_sheet(instance)
    inventory = build_inventory_template()

    item_clusters_sheet = ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET[table]

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        master.to_excel(
            writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False
        )
        regions_clusters.to_excel(
            writer,
            sheet_name=ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET,
            index=False,
        )
        items_clusters.to_excel(writer, sheet_name=item_clusters_sheet, index=False)
        db_units.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_DB_UNITS_SHEET, index=False)
        if redefine_uncertainties:
            build_uncertainties_sheet().to_excel(
                writer, sheet_name=ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET, index=False
            )
        for sheet_name in master["Inventory sheet"].tolist():
            inventory.to_excel(writer, sheet_name=sheet_name, index=False)
        if table == IOT:
            build_split_total_outputs_template().to_excel(
                writer, sheet_name=ADD_SECTOR_SPLIT_OUTPUT_SHEET, index=False
            )
            build_split_trades_template().to_excel(
                writer, sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET, index=False
            )
            build_split_exclusions_template().to_excel(
                writer, sheet_name=ADD_SECTOR_SPLIT_EXCLUSION_SHEET, index=False
            )
            build_split_tolerances_template().to_excel(
                writer, sheet_name=ADD_SECTOR_SPLIT_TOLERANCE_SHEET, index=False
            )


def write_inventory_templates_to_workbook(
    instance,
    workbook: AddSectorWorkbook,
    path: str | Path,
    *,
    overwrite: bool = True,
) -> list[str]:
    """Create missing inventory templates referenced by one add-sectors workbook.

    This mirrors the old ``get_inventory_sheets(...)`` workflow: it inspects the
    master sheet, creates one empty inventory tab per referenced sheet name,
    refreshes the ``DB units`` sheet, and for IOT workbooks also ensures the
    split-support tabs exist when at least one row is marked as ``Split``.

    Returns
    -------
    list[str]
        The inventory sheet names referenced by the workbook master sheet.
    """

    inventory_names = [
        sheet
        for sheet in workbook.master_sheet["Inventory sheet"].dropna().astype(str).tolist()
        if sheet
    ]
    inventory_names = list(dict.fromkeys(inventory_names))
    existing_sheets = set(load_workbook(path, data_only=False).sheetnames)
    inventory_template = build_inventory_template()
    db_units = build_db_units_sheet(instance)

    def _write_sheet(writer, sheet_name: str, frame: pd.DataFrame):
        if sheet_name in existing_sheets and not overwrite:
            return
        frame.to_excel(writer, sheet_name=sheet_name, index=False)

    split_sheets_needed = _workbook_has_split_rows(workbook)

    with pd.ExcelWriter(
        path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace",
    ) as writer:
        for sheet_name in inventory_names:
            _write_sheet(writer, sheet_name, inventory_template)

        _write_sheet(writer, ADVANCED_ADD_SECTOR_DB_UNITS_SHEET, db_units)

        if split_sheets_needed:
            _write_sheet(writer, ADD_SECTOR_SPLIT_OUTPUT_SHEET, build_split_total_outputs_template())
            _write_sheet(writer, ADD_SECTOR_SPLIT_TRADE_SHEET, build_split_trades_template())
            _write_sheet(writer, ADD_SECTOR_SPLIT_EXCLUSION_SHEET, build_split_exclusions_template())
            _write_sheet(writer, ADD_SECTOR_SPLIT_TOLERANCE_SHEET, build_split_tolerances_template())

    return inventory_names


def read_add_sector_workbook(
    path: str | Path,
    *,
    table: str,
    require_inventory_sheets: bool = False,
) -> AddSectorWorkbook:
    """Read and validate one add-sectors workbook.

    This function does not derive database-specific sets such as
    ``new_sectors`` or ``parented_activities``. It only parses the workbook
    structure into a normalized object. The database layer derives those sets
    later because they depend on the current database contents.

    When ``require_inventory_sheets`` is false, the workbook may reference
    inventory sheet names in the master sheet even if those sheets have not
    been created yet. This supports the workflow where users first fill the
    master sheet and only later add the inventory tabs. When it is true, all
    referenced inventory sheets must exist.
    """

    sheets = _read_excel_workbook(path)
    item_clusters_sheet = ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET[table]
    factors_clusters_sheet = ADVANCED_ADD_SECTOR_FACTORS_CLUSTERS_SHEET

    required = {
        ADVANCED_ADD_SECTOR_MASTER_SHEET,
        ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET,
        item_clusters_sheet,
    }
    missing = sorted(required.difference(sheets))
    if missing:
        raise WrongExcelFormat(
            f"Missing required sheets for add-sectors workbook: {missing}"
        )

    master_sheet = sheets[ADVANCED_ADD_SECTOR_MASTER_SHEET]
    expected_columns = list(ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[table].values())
    missing_columns = [column for column in expected_columns if column not in master_sheet.columns]
    if missing_columns:
        raise WrongExcelFormat(
            f"Add-sectors master sheet is missing columns: {missing_columns}"
        )
    master_sheet = master_sheet.loc[:, expected_columns]

    regions_clusters = _parse_cluster_sheet(
        sheets[ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET]
    )
    item_clusters = _parse_cluster_sheet(sheets[item_clusters_sheet])
    factors_clusters = _parse_cluster_sheet(sheets.get(factors_clusters_sheet, pd.DataFrame()))
    uncertainty_values = _parse_uncertainties_sheet(
        sheets.get(ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET)
    )

    inventory_names = [
        sheet
        for sheet in master_sheet["Inventory sheet"].dropna().astype(str).tolist()
        if sheet
    ]
    missing_inventories = [sheet for sheet in inventory_names if sheet not in sheets]
    if missing_inventories and require_inventory_sheets:
        raise WrongExcelFormat(
            f"Add-sectors workbook is missing inventory sheets: {missing_inventories}"
        )

    inventories_by_sheet = {
        # filter out missing sheets to allow partial workbooks
        sheet_name: sheets[sheet_name] for sheet_name in inventory_names
        if sheet_name in sheets
    }
    split_info = _parse_split_sheets(sheets) if _has_split_rows(table, master_sheet) else None

    return AddSectorWorkbook(
        table=table,
        master_sheet=master_sheet,
        regions_clusters=regions_clusters,
        item_clusters=item_clusters,
        factors_clusters=factors_clusters,
        uncertainty_values=uncertainty_values,
        inventories_by_sheet=inventories_by_sheet,
        split_info=split_info,
    )


def _read_excel_workbook(path: str | Path) -> dict[str, pd.DataFrame]:
    """Read one workbook while preserving cached Excel formulas when available.

    ``pandas.read_excel(...)`` reads formula cells through openpyxl's
    ``data_only`` path. When a workbook has not been recalculated by Excel yet,
    formula cells are therefore returned as ``NaN``. The add-sectors workflow
    needs a more robust behavior because inventories often contain formulas.

    This helper loads the workbook twice:
    - one view with formulas;
    - one view with cached values.

    It prefers cached values when present and otherwise evaluates a practical
    subset of Excel formulas directly from the formula text.
    """

    formula_book = load_workbook(path, data_only=False)
    values_book = load_workbook(path, data_only=True)
    cache: dict[tuple[str, str], Any] = {}
    sheets: dict[str, pd.DataFrame] = {}

    for sheet_name in formula_book.sheetnames:
        formula_sheet = formula_book[sheet_name]
        values_sheet = values_book[sheet_name]
        rows = []
        for row in formula_sheet.iter_rows():
            rows.append(
                [
                    _resolve_excel_cell(
                        sheet_name=sheet_name,
                        cell_coordinate=cell.coordinate,
                        formula_book=formula_book,
                        values_book=values_book,
                        cache=cache,
                        stack=set(),
                    )
                    for cell in row
                ]
            )

        rows = _trim_empty_excel_rows(rows)
        if not rows:
            sheets[sheet_name] = pd.DataFrame()
            continue

        header = rows[0]
        data = rows[1:]
        sheets[sheet_name] = pd.DataFrame(data, columns=header)

    return sheets


def _resolve_excel_cell(
    *,
    sheet_name: str,
    cell_coordinate: str,
    formula_book,
    values_book,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
):
    key = (sheet_name, cell_coordinate)
    if key in cache:
        return cache[key]
    if key in stack:
        raise WrongExcelFormat(f"Circular reference detected while reading formulas in {sheet_name}!{cell_coordinate}.")

    stack.add(key)
    formula_cell = formula_book[sheet_name][cell_coordinate]
    values_cell = values_book[sheet_name][cell_coordinate]
    raw_value = formula_cell.value
    cached_value = values_cell.value

    if isinstance(raw_value, str) and raw_value.startswith("="):
        if cached_value is not None:
            value = cached_value
        else:
            value = _evaluate_excel_formula(
                raw_value,
                current_sheet=sheet_name,
                formula_book=formula_book,
                values_book=values_book,
                cache=cache,
                stack=stack,
            )
    else:
        value = raw_value

    cache[key] = value
    stack.remove(key)
    return value


def _evaluate_excel_formula(
    formula: str,
    *,
    current_sheet: str,
    formula_book,
    values_book,
    cache: dict[tuple[str, str], Any],
    stack: set[tuple[str, str]],
):
    tokens = Tokenizer(formula).items
    parts: list[str] = []

    for token in tokens:
        if token.type == "FUNC":
            if token.subtype == "OPEN":
                parts.append(f"{token.value[:-1].upper()}(")
            else:
                parts.append(")")
        elif token.type == "OPERAND":
            if token.subtype == "RANGE":
                ref_sheet, ref_value = _split_excel_reference(token.value, current_sheet)
                if ":" in ref_value:
                    parts.append(f'_R("{ref_sheet}", "{ref_value}")')
                else:
                    parts.append(f'_C("{ref_sheet}", "{ref_value}")')
            elif token.subtype == "NUMBER":
                parts.append(token.value)
            elif token.subtype == "LOGICAL":
                parts.append("True" if token.value.upper() == "TRUE" else "False")
            elif token.subtype == "TEXT":
                parts.append(repr(token.value))
            else:
                parts.append(repr(token.value))
        elif token.type == "OPERATOR-INFIX":
            parts.append(
                {
                    "^": "**",
                    "=": "==",
                    "<>": "!=",
                    "&": "+",
                }.get(token.value, token.value)
            )
        elif token.type == "OPERATOR-PREFIX":
            parts.append(token.value)
        elif token.type == "OPERATOR-POSTFIX":
            continue
        elif token.type == "SEP":
            parts.append(",")
        elif token.type == "PAREN":
            parts.append(token.value)
        elif token.type == "WHITE-SPACE":
            continue
        else:
            raise WrongExcelFormat(f"Unsupported Excel token '{token.value}' while reading add-sectors formulas.")

    expression = "".join(parts)

    def _cell(sheet: str, ref: str):
        value = _resolve_excel_cell(
            sheet_name=sheet,
            cell_coordinate=ref,
            formula_book=formula_book,
            values_book=values_book,
            cache=cache,
            stack=stack,
        )
        return 0 if value in (None, "") else value

    def _range(sheet: str, ref: str):
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        values = []
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                coord = f"{_column_name(col)}{row}"
                values.append(_cell(sheet, coord))
        return values

    env = {
        "_C": _cell,
        "_R": _range,
        "SUM": lambda *args: sum(_flatten_formula_args(args)),
        "AVERAGE": lambda *args: _average_formula_args(args),
        "MIN": lambda *args: min(_flatten_formula_args(args)),
        "MAX": lambda *args: max(_flatten_formula_args(args)),
        "ABS": abs,
        "ROUND": round,
        "IF": lambda cond, yes, no: yes if cond else no,
        "AND": lambda *args: all(_flatten_formula_args(args)),
        "OR": lambda *args: any(_flatten_formula_args(args)),
        "NOT": lambda arg: not arg,
    }

    try:
        return eval(expression, {"__builtins__": {}}, env)
    except Exception as exc:
        raise WrongExcelFormat(
            f"Could not evaluate Excel formula '{formula}' in add-sectors workbook."
        ) from exc


def _split_excel_reference(reference: str, current_sheet: str) -> tuple[str, str]:
    if "!" in reference:
        sheet_name, cell_ref = reference.split("!", 1)
        sheet_name = sheet_name.strip("'")
    else:
        sheet_name = current_sheet
        cell_ref = reference
    return sheet_name, cell_ref.replace("$", "")


def _trim_empty_excel_rows(rows: list[list[Any]]) -> list[list[Any]]:
    def _is_empty(value: Any) -> bool:
        return value is None or value == ""

    max_row = 0
    max_col = 0
    for row_index, row in enumerate(rows, start=1):
        if any(not _is_empty(value) for value in row):
            max_row = row_index
        for col_index, value in enumerate(row, start=1):
            if not _is_empty(value):
                max_col = max(max_col, col_index)

    if max_row == 0 or max_col == 0:
        return []

    return [row[:max_col] for row in rows[:max_row]]


def _flatten_formula_args(args) -> list[Any]:
    values: list[Any] = []
    for arg in args:
        if isinstance(arg, list):
            values.extend(_flatten_formula_args(arg))
        else:
            values.append(0 if arg in (None, "") else arg)
    return values


def _average_formula_args(args) -> float:
    values = _flatten_formula_args(args)
    return sum(values) / len(values) if values else 0


def _column_name(index: int) -> str:
    result = []
    while index:
        index, remainder = divmod(index - 1, 26)
        result.append(chr(65 + remainder))
    return "".join(reversed(result))


def derive_add_sector_sets(
    workbook: AddSectorWorkbook,
    *,
    existing_sectors: list[str] | None = None,
    existing_activities: list[str] | None = None,
    existing_commodities: list[str] | None = None,
) -> dict[str, list[str]]:
    """Derive the item sets consumed by the add-sectors engine.

    The workbook stores rows, not semantic groups. This helper turns the raw
    master sheet into the derived sets used by ``Database.add_sectors(...)``,
    such as ``new_sectors``, ``new_activities``, ``parented_*`` and
    ``to_split_sectors``.
    """

    master = workbook.master_sheet
    if workbook.table == IOT:
        if existing_sectors is None:
            raise WrongInput("existing_sectors is required to derive IOT add-sector sets.")
        sectors = [
            sector
            for sector in master[_MASTER_INDEX["s"]].dropna().astype(str).unique().tolist()
            if sector not in existing_sectors
        ]
        parented = []
        parent_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[IOT]["parent_sector"]
        for sector in sectors:
            parent = master.loc[master[_MASTER_INDEX["s"]] == sector, parent_column].iloc[0]
            if isinstance(parent, str) and parent.strip():
                parented.append(sector)
        split_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[IOT]["add_mode"]
        to_split = [
            sector
            for sector in master.loc[
                master[split_column].fillna("").astype(str).str.lower() == "split",
                _MASTER_INDEX["s"],
            ]
            .astype(str)
            .unique()
            .tolist()
            if sector in sectors
        ]
        non_parented = [sector for sector in sectors if sector not in parented]
        return {
            "new_sectors": sectors,
            "parented_sectors": parented,
            "non_parented_sectors": non_parented,
            "to_split_sectors": to_split,
        }

    if existing_activities is None or existing_commodities is None:
        raise WrongInput(
            "existing_activities and existing_commodities are required to derive SUT add-sector sets."
        )

    activities = [
        value
        for value in master[_MASTER_INDEX["a"]].dropna().astype(str).unique().tolist()
        if value and value not in existing_activities
    ]
    commodities = [
        value
        for value in master[_MASTER_INDEX["c"]].dropna().astype(str).unique().tolist()
        if value and value not in existing_commodities
    ]
    parented = []
    parent_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[SUT]["parent_activity"]
    for activity in activities:
        parent = master.loc[master[_MASTER_INDEX["a"]] == activity, parent_column].iloc[0]
        if isinstance(parent, str) and parent.strip():
            parented.append(activity)
    non_parented = [activity for activity in activities if activity not in parented]
    return {
        "new_activities": activities,
        "new_commodities": commodities,
        "parented_activities": parented,
        "non_parented_activities": non_parented,
    }


def _parse_cluster_sheet(sheet: pd.DataFrame) -> dict[str, list[str]]:
    clusters: dict[str, list[str]] = {}
    for column in sheet.columns:
        values = [str(value) for value in sheet[column].dropna().tolist() if str(value).strip()]
        if values:
            clusters[str(column)] = values
    return clusters


def _parse_uncertainties_sheet(sheet: pd.DataFrame | None) -> dict[str, float]:
    if sheet is None:
        return dict(ADVANCED_ADD_SECTOR_UNCERTAINTY_PARAMETERS)
    if list(sheet.columns[:2]) != ADVANCED_ADD_SECTOR_UNCERTAINTY_COLUMNS:
        raise WrongExcelFormat(
            "Add-sectors uncertainties sheet should expose the expected two columns."
        )
    return dict(zip(sheet.iloc[:, 0], sheet.iloc[:, 1]))


def _parse_split_sheet(
    sheets: dict[str, pd.DataFrame],
    *,
    sheet_name: str,
    columns: dict[str, str],
) -> pd.DataFrame:
    """Return one normalized split-support sheet."""

    if sheet_name not in sheets:
        raise WrongExcelFormat(
            f"Missing required split-support sheet in add-sectors workbook: {sheet_name}"
        )
    sheet = sheets[sheet_name].copy()
    expected = list(columns.values())
    missing_columns = [column for column in expected if column not in sheet.columns]
    if missing_columns:
        raise WrongExcelFormat(
            f"Split-support sheet '{sheet_name}' is missing columns: {missing_columns}"
        )
    return sheet.loc[:, expected]


def _parse_split_sheets(sheets: dict[str, pd.DataFrame]) -> dict[str, pd.DataFrame]:
    """Return the normalized split-support payload for IOT workbooks."""

    return {
        ADD_SECTOR_SPLIT_OUTPUT_SHEET: _parse_split_sheet(
            sheets,
            sheet_name=ADD_SECTOR_SPLIT_OUTPUT_SHEET,
            columns=ADD_SECTOR_SPLIT_OUTPUT_COLUMNS,
        ),
        ADD_SECTOR_SPLIT_TRADE_SHEET: _parse_split_sheet(
            sheets,
            sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET,
            columns=ADD_SECTOR_SPLIT_TRADE_COLUMNS,
        ),
        ADD_SECTOR_SPLIT_EXCLUSION_SHEET: _parse_split_sheet(
            sheets,
            sheet_name=ADD_SECTOR_SPLIT_EXCLUSION_SHEET,
            columns=ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS,
        ),
        ADD_SECTOR_SPLIT_TOLERANCE_SHEET: _parse_split_sheet(
            sheets,
            sheet_name=ADD_SECTOR_SPLIT_TOLERANCE_SHEET,
            columns=ADD_SECTOR_SPLIT_TOLERANCE_COLUMNS,
        ),
    }


def _has_split_rows(table: str, master_sheet: pd.DataFrame) -> bool:
    """Return whether an IOT add-sectors workbook contains split rows."""
    if table != IOT:
        return False
    add_mode_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[IOT]["add_mode"]
    if add_mode_column not in master_sheet.columns:
        return False
    values = (
        master_sheet[add_mode_column]
        .dropna()
        .astype(str)
        .str.strip()
        .str.lower()
    )
    return values.eq("split").any()


def _workbook_has_split_rows(workbook: AddSectorWorkbook) -> bool:
    """Return whether an IOT add-sectors workbook contains split rows."""
    return _has_split_rows(workbook.table, workbook.master_sheet)


def group_inventories_by_target(
    workbook: AddSectorWorkbook,
) -> dict[str, dict[str, pd.DataFrame]]:
    """Group inventory sheets by target item name.

    The engine consumes inventories one target item at a time. This helper maps
    each target item name to the inventory sheets referenced by that item in the
    master sheet.
    """

    master = workbook.master_sheet
    table = workbook.table
    inventory_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[table]["inventory_sheet"]
    target_column = _MASTER_INDEX["a"] if table == SUT else _MASTER_INDEX["s"]

    grouped: dict[str, dict[str, pd.DataFrame]] = {}
    for _, row in master.iterrows():
        sheet_name = row[inventory_column]
        target = row[target_column]
        if pd.isna(sheet_name) or pd.isna(target):
            continue
        sheet_name = str(sheet_name)
        target = str(target)
        inventory = workbook.inventories_by_sheet.get(sheet_name)
        if inventory is None:
            continue
        grouped.setdefault(target, {})[sheet_name] = inventory

    return grouped


def _copy_add_sector_workbook(workbook: AddSectorWorkbook) -> AddSectorWorkbook:
    return AddSectorWorkbook(
        table=workbook.table,
        master_sheet=workbook.master_sheet.copy(deep=True),
        regions_clusters={key: list(values) for key, values in workbook.regions_clusters.items()},
        item_clusters={key: list(values) for key, values in workbook.item_clusters.items()},
        factors_clusters={key: list(values) for key, values in workbook.factors_clusters.items()},
        uncertainty_values=dict(workbook.uncertainty_values),
        inventories_by_sheet={
            sheet_name: frame.copy(deep=True)
            for sheet_name, frame in workbook.inventories_by_sheet.items()
        },
        split_info=(
            {sheet_name: frame.copy(deep=True) for sheet_name, frame in workbook.split_info.items()}
            if workbook.split_info
            else None
        ),
    )


def _unique_non_empty_values(frame: pd.DataFrame, column: str) -> list[str]:
    if column not in frame.columns:
        return []

    values: list[str] = []
    for value in frame[column].dropna().tolist():
        text = str(value).strip()
        if text and text not in values:
            values.append(text)

    return values


def _cluster_signature(values: list[str]) -> tuple[str, ...]:
    normalized = {str(value).strip() for value in values if str(value).strip()}
    return tuple(sorted(normalized))


def _next_progressive_name(base_name: str, used_names: set[str]) -> str:
    counter = 1
    while True:
        candidate = f"{base_name} {counter}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        counter += 1


def _rename_dataframe_column_values(
    frame: pd.DataFrame,
    column: str,
    rename_map: dict[str, str],
    *,
    row_mask: pd.Series | None = None,
) -> None:
    if column not in frame.columns or not rename_map:
        return

    mask = frame[column].notna()
    if row_mask is not None:
        mask &= row_mask

    if not mask.any():
        return

    values = frame.loc[mask, column].astype(str)
    replace_mask = values.isin(rename_map)
    if not replace_mask.any():
        return

    frame.loc[values.index[replace_mask], column] = values.loc[replace_mask].map(rename_map).values


def _rename_inventory_db_item_values(
    inventories_by_sheet: dict[str, pd.DataFrame],
    rename_map: dict[str, str],
    *,
    allowed_item_type: str,
) -> None:
    item_type_column = ADVANCED_ADD_SECTOR_INVENTORY_SHEET_COLUMNS["item_type"]
    db_item_column = ADVANCED_ADD_SECTOR_INVENTORY_SHEET_COLUMNS["db_item"]

    for frame in inventories_by_sheet.values():
        if item_type_column not in frame.columns or db_item_column not in frame.columns:
            continue
        row_mask = frame[item_type_column].fillna("").astype(str).eq(allowed_item_type)
        _rename_dataframe_column_values(frame, db_item_column, rename_map, row_mask=row_mask)


def _rename_inventory_db_region_values(
    inventories_by_sheet: dict[str, pd.DataFrame],
    rename_map: dict[str, str],
) -> None:
    db_region_column = ADVANCED_ADD_SECTOR_INVENTORY_SHEET_COLUMNS["db_region"]

    for frame in inventories_by_sheet.values():
        _rename_dataframe_column_values(frame, db_region_column, rename_map)


def _rename_split_info_region_values(
    split_info: dict[str, pd.DataFrame] | None,
    rename_map: dict[str, str],
) -> None:
    if not split_info:
        return

    _rename_dataframe_column_values(
        split_info[ADD_SECTOR_SPLIT_OUTPUT_SHEET],
        ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["region"],
        rename_map,
    )
    _rename_dataframe_column_values(
        split_info[ADD_SECTOR_SPLIT_TRADE_SHEET],
        ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_from"],
        rename_map,
    )
    _rename_dataframe_column_values(
        split_info[ADD_SECTOR_SPLIT_TRADE_SHEET],
        ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_to"],
        rename_map,
    )
    _rename_dataframe_column_values(
        split_info[ADD_SECTOR_SPLIT_EXCLUSION_SHEET],
        ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS["region_from"],
        rename_map,
    )
    _rename_dataframe_column_values(
        split_info[ADD_SECTOR_SPLIT_EXCLUSION_SHEET],
        ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS["region_to"],
        rename_map,
    )


def _validate_distinct_master_targets(workbooks: list[AddSectorWorkbook]) -> None:
    if not workbooks:
        raise WrongInput("At least one add-sectors workbook is required.")

    table = workbooks[0].table

    def duplicates_for(column: str) -> list[str]:
        owners: dict[str, set[int]] = {}
        for workbook_index, workbook in enumerate(workbooks):
            for value in _unique_non_empty_values(workbook.master_sheet, column):
                owners.setdefault(value, set()).add(workbook_index)
        return sorted(value for value, workbook_indexes in owners.items() if len(workbook_indexes) > 1)

    if table == IOT:
        duplicates = duplicates_for(_MASTER_INDEX["s"])
        if duplicates:
            raise WrongInput(
                "Add-sectors workbooks define duplicate sectors across files: "
                f"{duplicates}"
            )
        return

    activity_duplicates = duplicates_for(_MASTER_INDEX["a"])
    if activity_duplicates:
        raise WrongInput(
            "Add-sectors workbooks define duplicate activities across files: "
            f"{activity_duplicates}"
        )

    commodity_duplicates = duplicates_for(_MASTER_INDEX["c"])
    if commodity_duplicates:
        raise WrongInput(
            "Add-sectors workbooks define duplicate commodities across files: "
            f"{commodity_duplicates}"
        )


def _rename_conflicting_inventory_sheets(workbooks: list[AddSectorWorkbook]) -> None:
    inventory_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[workbooks[0].table]["inventory_sheet"]
    occurrences: dict[str, list[int]] = {}

    for workbook_index, workbook in enumerate(workbooks):
        for sheet_name in _unique_non_empty_values(workbook.master_sheet, inventory_column):
            occurrences.setdefault(sheet_name, []).append(workbook_index)

    used_names = {
        sheet_name for sheet_name, workbook_indexes in occurrences.items() if len(workbook_indexes) == 1
    }
    rename_maps = [dict() for _ in workbooks]

    for sheet_name, workbook_indexes in occurrences.items():
        if len(workbook_indexes) == 1:
            rename_maps[workbook_indexes[0]][sheet_name] = sheet_name
            continue

        for workbook_index in workbook_indexes:
            rename_maps[workbook_index][sheet_name] = _next_progressive_name(sheet_name, used_names)

    for workbook, rename_map in zip(workbooks, rename_maps):
        rename_map = {old: new for old, new in rename_map.items() if old != new}
        if not rename_map:
            continue

        _rename_dataframe_column_values(workbook.master_sheet, inventory_column, rename_map)
        renamed = {
            rename_map.get(sheet_name, sheet_name): frame.copy(deep=True)
            for sheet_name, frame in workbook.inventories_by_sheet.items()
        }
        workbook.inventories_by_sheet.clear()
        workbook.inventories_by_sheet.update(renamed)


def _rename_region_cluster_references(
    workbook: AddSectorWorkbook,
    rename_map: dict[str, str],
) -> None:
    region_column = ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS[workbook.table]["region"]
    _rename_dataframe_column_values(workbook.master_sheet, region_column, rename_map)
    _rename_inventory_db_region_values(workbook.inventories_by_sheet, rename_map)
    _rename_split_info_region_values(workbook.split_info, rename_map)


def _rename_item_cluster_references(
    workbook: AddSectorWorkbook,
    rename_map: dict[str, str],
) -> None:
    allowed_item_type = _MASTER_INDEX["c"] if workbook.table == SUT else _MASTER_INDEX["s"]
    _rename_inventory_db_item_values(
        workbook.inventories_by_sheet,
        rename_map,
        allowed_item_type=allowed_item_type,
    )


def _rename_factor_cluster_references(
    workbook: AddSectorWorkbook,
    rename_map: dict[str, str],
) -> None:
    _rename_inventory_db_item_values(
        workbook.inventories_by_sheet,
        rename_map,
        allowed_item_type=_MASTER_INDEX["f"],
    )


def _rename_conflicting_cluster_names(
    workbooks: list[AddSectorWorkbook],
    *,
    attr_name: str,
    reference_updater,
) -> None:
    occurrences: dict[str, list[tuple[int, tuple[str, ...]]]] = {}

    for workbook_index, workbook in enumerate(workbooks):
        cluster_map = getattr(workbook, attr_name)
        for cluster_name, members in cluster_map.items():
            name = str(cluster_name).strip()
            if not name:
                continue
            occurrences.setdefault(name, []).append((workbook_index, _cluster_signature(members)))

    used_names = {
        name
        for name, cluster_occurrences in occurrences.items()
        if len({signature for _, signature in cluster_occurrences}) == 1
    }
    rename_maps = [dict() for _ in workbooks]

    for cluster_name, cluster_occurrences in occurrences.items():
        distinct_signatures = {signature for _, signature in cluster_occurrences}
        if len(distinct_signatures) == 1:
            for workbook_index, _ in cluster_occurrences:
                rename_maps[workbook_index][cluster_name] = cluster_name
            continue

        signature_to_name: dict[tuple[str, ...], str] = {}
        for workbook_index, signature in cluster_occurrences:
            final_name = signature_to_name.get(signature)
            if final_name is None:
                final_name = _next_progressive_name(cluster_name, used_names)
                signature_to_name[signature] = final_name
            rename_maps[workbook_index][cluster_name] = final_name

    for workbook, rename_map in zip(workbooks, rename_maps):
        rename_map = {old: new for old, new in rename_map.items() if old != new}
        if not rename_map:
            continue

        cluster_map = getattr(workbook, attr_name)
        renamed = {
            rename_map.get(cluster_name, cluster_name): list(members)
            for cluster_name, members in cluster_map.items()
        }
        cluster_map.clear()
        cluster_map.update(renamed)
        reference_updater(workbook, rename_map)


def _merge_cluster_maps(
    workbooks: list[AddSectorWorkbook],
    *,
    attr_name: str,
) -> dict[str, list[str]]:
    merged: dict[str, list[str]] = {}

    for workbook in workbooks:
        cluster_map = getattr(workbook, attr_name)
        for cluster_name, members in cluster_map.items():
            values = [str(value).strip() for value in members if str(value).strip()]
            if cluster_name not in merged:
                merged[cluster_name] = values
                continue

            if _cluster_signature(merged[cluster_name]) != _cluster_signature(values):
                raise WrongInput(
                    f"Add-sectors cluster '{cluster_name}' has conflicting definitions across files."
                )

    return merged


def _merge_uncertainty_values(workbooks: list[AddSectorWorkbook]) -> dict[str, float]:
    merged: dict[str, float] = {}

    for workbook in workbooks:
        for key, value in workbook.uncertainty_values.items():
            if key not in merged:
                merged[key] = value
                continue

            if merged[key] != value:
                raise WrongInput(
                    f"Add-sectors uncertainty '{key}' has conflicting values across files."
                )

    return merged


def _merge_split_info(
    workbooks: list[AddSectorWorkbook],
) -> dict[str, pd.DataFrame] | None:
    split_payloads = [workbook.split_info for workbook in workbooks if workbook.split_info]
    if not split_payloads:
        return None

    merged: dict[str, pd.DataFrame] = {}
    for sheet_name in split_payloads[0]:
        frames = [payload[sheet_name].copy(deep=True) for payload in split_payloads]
        if sheet_name == ADD_SECTOR_SPLIT_TOLERANCE_SHEET:
            reference = frames[0].reset_index(drop=True)
            for frame in frames[1:]:
                try:
                    pd.testing.assert_frame_equal(
                        reference,
                        frame.reset_index(drop=True),
                        check_dtype=False,
                        check_like=True,
                    )
                except AssertionError as exc:
                    raise WrongInput(
                        "Add-sectors workbooks define conflicting split tolerance sheets."
                    ) from exc
            merged[sheet_name] = reference
            continue

        merged[sheet_name] = pd.concat(frames, ignore_index=True)

    return merged


def merge_add_sector_workbooks(
    workbooks: list[AddSectorWorkbook],
) -> AddSectorWorkbook:
    """Merge multiple normalized add-sectors workbooks into one payload."""

    if not workbooks:
        raise WrongInput("At least one add-sectors workbook is required.")

    copied = [_copy_add_sector_workbook(workbook) for workbook in workbooks]
    table = copied[0].table
    if any(workbook.table != table for workbook in copied):
        raise WrongInput("All add-sectors workbooks must target the same table type.")

    _validate_distinct_master_targets(copied)
    _rename_conflicting_inventory_sheets(copied)
    _rename_conflicting_cluster_names(
        copied,
        attr_name="regions_clusters",
        reference_updater=_rename_region_cluster_references,
    )
    _rename_conflicting_cluster_names(
        copied,
        attr_name="item_clusters",
        reference_updater=_rename_item_cluster_references,
    )
    _rename_conflicting_cluster_names(
        copied,
        attr_name="factors_clusters",
        reference_updater=_rename_factor_cluster_references,
    )

    return AddSectorWorkbook(
        table=table,
        master_sheet=pd.concat([workbook.master_sheet for workbook in copied], ignore_index=True),
        regions_clusters=_merge_cluster_maps(copied, attr_name="regions_clusters"),
        item_clusters=_merge_cluster_maps(copied, attr_name="item_clusters"),
        factors_clusters=_merge_cluster_maps(copied, attr_name="factors_clusters"),
        uncertainty_values=_merge_uncertainty_values(copied),
        inventories_by_sheet={
            sheet_name: frame.copy(deep=True)
            for workbook in copied
            for sheet_name, frame in workbook.inventories_by_sheet.items()
        },
        split_info=_merge_split_info(copied),
    )


# Backward-compatible aliases for the first port of the workflow.
AdvancedAddSectorWorkbook = AddSectorWorkbook
build_advanced_master_sheet = build_add_sector_master_sheet
write_advanced_add_sector_workbook = write_add_sector_workbook
read_advanced_add_sector_workbook = read_add_sector_workbook
derive_advanced_add_sector_sets = derive_add_sector_sets
group_advanced_inventories_by_target = group_inventories_by_target
merge_advanced_add_sector_workbooks = merge_add_sector_workbooks
