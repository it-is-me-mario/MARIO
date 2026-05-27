import warnings
warnings.filterwarnings("ignore",category=DeprecationWarning)

import inspect
import hashlib
from pathlib import Path

from openpyxl import load_workbook

from mario.ops.add_sector_specs import (
    ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS,
    ADD_SECTOR_SPLIT_EXCLUSION_SHEET,
    ADD_SECTOR_SPLIT_OUTPUT_COLUMNS,
    ADD_SECTOR_SPLIT_OUTPUT_SHEET,
    ADD_SECTOR_SPLIT_TOLERANCE_COLUMNS,
    ADD_SECTOR_SPLIT_TOLERANCE_DEFAULTS,
    ADD_SECTOR_SPLIT_TOLERANCE_SHEET,
    ADD_SECTOR_SPLIT_TRADE_COLUMNS,
    ADD_SECTOR_SPLIT_TRADE_SHEET,
    ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET,
    ADVANCED_ADD_SECTOR_MASTER_SHEET,
    ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS,
    ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET,
    ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET,
)
from mario.ops import cvxlab_bridge as bridge_module
from mario.ops.cvxlab_bridge import (
    CVXLAB_SPLIT_MODEL_NAME,
    _build_cvxlab_model,
    _load_model_coordinates_and_initialize_database,
)
from mario.ops.add_sector_workbook import (
    derive_advanced_add_sector_sets,
    group_advanced_inventories_by_target,
    read_advanced_add_sector_workbook,
)
from mario.log_exc.exceptions import NotImplementable, WrongExcelFormat, WrongInput
from mario.ops.sectoradd import get_corresponding_keys,matrix_concat,fill_matrix
from mario.model.conventions import _MASTER_INDEX
from mario.model.builders import MatrixBuilder
from mario.ops.workbook_specs import ADD_SECTOR_SHEETS
import pandas.testing as pdt
import pandas as pd
import pytest
from mario.api import database as database_module
from mario import load_test
from mario.parsers.api import build_database_from_state, build_parser_state
from mario.parsers.entrypoints import parse_from_excel
from mario.parsers.matrix_layouts import sut_block_specs_for_matrix_layouts
from mario.model.conventions import _ENUM,_MASTER_INDEX

@pytest.fixture()
def CoreDataIOT():
    return load_test("IOT")


@pytest.fixture()
def CoreDataSUT():
    return load_test("SUT")


def _configure_split_workbook(instance, path, *, new_sector="Split sector", quantity=0.1):
    region = instance.get_index(_MASTER_INDEX["r"])[0]
    other_region = (
        instance.get_index(_MASTER_INDEX["r"])[1]
        if len(instance.get_index(_MASTER_INDEX["r"])) > 1
        else region
    )
    parent_sector = instance.get_index(_MASTER_INDEX["s"])[0]
    unit = instance.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    master.loc[0, "Add or Split"] = "Split"

    split_outputs = pd.DataFrame(
        [
            {
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["sector"]: new_sector,
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["region"]: region,
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["quantity"]: quantity,
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["unit"]: unit,
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["source"]: "test",
                ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["notes"]: "",
            }
        ]
    )
    split_trades = pd.DataFrame(
        [
            {
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["sector_from"]: new_sector,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_from"]: region,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_to"]: other_region,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["quantity"]: quantity,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["unit"]: unit,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["source"]: "test",
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["notes"]: "",
            }
        ]
    )
    split_exclusions = pd.DataFrame(columns=list(ADD_SECTOR_SPLIT_EXCLUSION_COLUMNS.values()))
    split_tolerances = pd.DataFrame(
        ADD_SECTOR_SPLIT_TOLERANCE_DEFAULTS,
        columns=list(ADD_SECTOR_SPLIT_TOLERANCE_COLUMNS.values()),
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        split_outputs.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_OUTPUT_SHEET, index=False)
        split_trades.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET, index=False)
        split_exclusions.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_EXCLUSION_SHEET, index=False)
        split_tolerances.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_TOLERANCE_SHEET, index=False)

    return {
        "region": region,
        "other_region": other_region,
        "parent_sector": parent_sector,
        "unit": unit,
        "new_sector": new_sector,
    }


def _configure_directory_add_sector_workbook(
    instance,
    path,
    *,
    new_sector,
    target_region,
    inventory_sheet,
    cluster_name=None,
    cluster_members=None,
):
    instance.get_add_sectors_excel(items=[new_sector], regions=[target_region], path=path)

    unit = instance.units[_MASTER_INDEX["s"]].iloc[0, 0]
    db_item = instance.get_index(_MASTER_INDEX["s"])[0]
    inventory_region = target_region

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[:, "Inventory sheet"] = inventory_sheet

    sheets_to_write = {
        ADVANCED_ADD_SECTOR_MASTER_SHEET: master,
    }

    if cluster_name is not None:
        inventory_region = cluster_name
        master.loc[:, _MASTER_INDEX["r"]] = cluster_name
        sheets_to_write[ADVANCED_ADD_SECTOR_MASTER_SHEET] = master
        sheets_to_write[ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET] = pd.DataFrame(
            {cluster_name: cluster_members or [target_region]}
        )

    sheets_to_write[inventory_sheet] = pd.DataFrame(
        [
            {
                "Quantity": 1,
                "Unit": unit,
                "Input": "demo input",
                "Item type": _MASTER_INDEX["s"],
                "DB Item": db_item,
                "DB Region": inventory_region,
                "Change type": "Update",
                "Source": "test",
                "Notes": "note",
            }
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        for sheet_name, frame in sheets_to_write.items():
            frame.to_excel(writer, sheet_name=sheet_name, index=False)


def _write_legacy_regional_extension_iot(path):
    flows = pd.DataFrame(
        [
            [None, None, None, "r1", "r1", "r2", "r2", "r1", "r2"],
            [None, None, None, "Sector", "Sector", "Sector", "Sector", "Consumption category", "Consumption category"],
            [None, None, None, "s1", "s2", "s1", "s2", "CC", "CC"],
            ["r1", "Sector", "s1", 20, 15, 4, 5, 15, 25],
            ["r1", "Sector", "s2", 10, 54, 5, 5, 20, 45],
            ["r2", "Sector", "s1", 5, 5, 18, 11, 15, 30],
            ["r2", "Sector", "s2", 3, 5, 9, 41, 35, 30],
            [None, "Factor of production", "VA", 46, 60, 48, 61, 0, 0],
            ["r1", "Satellite account", "CO2", 10, 5, 20, 10, 4, 0],
            ["r2", "Satellite account", "CO2", 23, 6, 5, 2, 1, 2],
        ]
    )
    units = pd.DataFrame(
        [
            [None, None, "unit"],
            ["Sector", "s1", "EUR"],
            ["Sector", "s2", "EUR"],
            ["Factor of production", "VA", "EUR"],
            ["Satellite account", "CO2", "ton"],
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        flows.to_excel(writer, sheet_name="flows", header=False, index=False)
        units.to_excel(writer, sheet_name="units", header=False, index=False)


def _write_explicit_layout_iot(path):
    flows = pd.DataFrame(
        [
            [None, None, None, "r1", "r1", "r2", "r2", None, None],
            [None, None, None, "s1", "s2", "s1", "s2", "hh", "investment"],
            ["r1", "s1", None, 20, 15, 4, 5, 15, 25],
            ["r1", "s2", None, 10, 54, 5, 5, 20, 45],
            ["r2", "s1", None, 5, 5, 18, 11, 15, 30],
            ["r2", "s2", None, 3, 5, 9, 41, 35, 30],
            ["r1", "s1", "taxes", 10, 5, 2, 1, 0, 0],
            ["r1", "s1", "capital", 18, 4, 1, 1, 0, 0],
            ["r1", "s2", "taxes", 1, 21, 4, 2, 0, 0],
            ["r1", "s2", "capital", 5, 15, 2, 3, 0, 0],
            ["r2", "s1", "taxes", 2, 0, 12, 7, 0, 0],
            ["r2", "s1", "capital", 1, 0, 8, 10, 0, 0],
            ["r2", "s2", "taxes", 3, 2, 2, 32, 0, 0],
            ["r2", "s2", "capital", 2, 4, 6, 18, 0, 0],
            ["r1", "CO2", None, 10, 5, 20, 10, 4, 0],
            ["r2", "CO2", None, 23, 6, 5, 2, 1, 2],
        ]
    )
    units = pd.DataFrame(
        [
            [None, None, "unit"],
            ["Sector", "s1", "EUR"],
            ["Sector", "s2", "EUR"],
            ["Factor of production", "taxes", "EUR"],
            ["Factor of production", "capital", "EUR"],
            ["Satellite account", "CO2", "ton"],
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        flows.to_excel(writer, sheet_name="flows", header=False, index=False)
        units.to_excel(writer, sheet_name="units", header=False, index=False)


def _write_explicit_layout_iot_with_regional_final_demand(path):
    flows = pd.DataFrame(
        [
            [None, None, None, "r1", "r1", "r2", "r2", "r1", "r2"],
            [None, None, None, "s1", "s2", "s1", "s2", "hh", "investment"],
            ["r1", "s1", None, 20, 15, 4, 5, 15, 25],
            ["r1", "s2", None, 10, 54, 5, 5, 20, 45],
            ["r2", "s1", None, 5, 5, 18, 11, 15, 30],
            ["r2", "s2", None, 3, 5, 9, 41, 35, 30],
            ["r1", "s1", "taxes", 10, 5, 2, 1, 0, 0],
            ["r1", "s1", "capital", 18, 4, 1, 1, 0, 0],
            ["r1", "s2", "taxes", 1, 21, 4, 2, 0, 0],
            ["r1", "s2", "capital", 5, 15, 2, 3, 0, 0],
            ["r2", "s1", "taxes", 2, 0, 12, 7, 0, 0],
            ["r2", "s1", "capital", 1, 0, 8, 10, 0, 0],
            ["r2", "s2", "taxes", 3, 2, 2, 32, 0, 0],
            ["r2", "s2", "capital", 2, 4, 6, 18, 0, 0],
            ["r1", "CO2", None, 10, 5, 20, 10, 4, 0],
            ["r2", "CO2", None, 23, 6, 5, 2, 1, 2],
        ]
    )
    units = pd.DataFrame(
        [
            [None, None, "unit"],
            ["Sector", "s1", "EUR"],
            ["Sector", "s2", "EUR"],
            ["Factor of production", "taxes", "EUR"],
            ["Factor of production", "capital", "EUR"],
            ["Satellite account", "CO2", "ton"],
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        flows.to_excel(writer, sheet_name="flows", header=False, index=False)
        units.to_excel(writer, sheet_name="units", header=False, index=False)


def _write_legacy_layout_iot_with_regional_factors(path):
    flows = pd.DataFrame(
        [
            [None, None, None, None, "r1", "r1", "r2", "r2", "r1", "r2"],
            [None, None, None, None, "Sector", "Sector", "Sector", "Sector", "Consumption category", "Consumption category"],
            [None, None, None, None, "s1", "s2", "s1", "s2", "CC", "CC"],
            ["r1", "Sector", "s1", None, 20, 15, 4, 5, 15, 25],
            ["r1", "Sector", "s2", None, 10, 54, 5, 5, 20, 45],
            ["r2", "Sector", "s1", None, 5, 5, 18, 11, 15, 30],
            ["r2", "Sector", "s2", None, 3, 5, 9, 41, 35, 30],
            ["r1", "s1", "Factor of production", "taxes", 10, 5, 2, 1, 0, 0],
            ["r1", "s1", "Factor of production", "capital", 18, 4, 1, 1, 0, 0],
            ["r1", "s2", "Factor of production", "taxes", 1, 21, 4, 2, 0, 0],
            ["r1", "s2", "Factor of production", "capital", 5, 15, 2, 3, 0, 0],
            ["r2", "s1", "Factor of production", "taxes", 2, 0, 12, 7, 0, 0],
            ["r2", "s1", "Factor of production", "capital", 1, 0, 8, 10, 0, 0],
            ["r2", "s2", "Factor of production", "taxes", 3, 2, 2, 32, 0, 0],
            ["r2", "s2", "Factor of production", "capital", 2, 4, 6, 18, 0, 0],
            ["r1", "Satellite account", "CO2", None, 10, 5, 20, 10, 4, 0],
            ["r2", "Satellite account", "CO2", None, 23, 6, 5, 2, 1, 2],
        ]
    )
    units = pd.DataFrame(
        [
            [None, None, "unit"],
            ["Sector", "s1", "EUR"],
            ["Sector", "s2", "EUR"],
            ["Factor of production", "taxes", "EUR"],
            ["Factor of production", "capital", "EUR"],
            ["Satellite account", "CO2", "ton"],
        ]
    )
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        flows.to_excel(writer, sheet_name="flows", header=False, index=False)
        units.to_excel(writer, sheet_name="units", header=False, index=False)


def _build_custom_sut_database():
    productive = pd.MultiIndex.from_tuples(
        [
            ("r1", "Activity", "a1"),
            ("r1", "Activity", "a2"),
            ("r1", "Commodity", "c1"),
            ("r1", "Commodity", "c2"),
        ],
        names=["Region", "Level", "Item"],
    )
    final_demand = pd.MultiIndex.from_tuples(
        [("r1", "Consumption category", "hh")],
        names=["Region", "Level", "Item"],
    )
    factors = pd.MultiIndex.from_tuples(
        [
            ("r1", "a1", "taxes"),
            ("r1", "a1", "capital"),
            ("r1", "a2", "taxes"),
            ("r1", "a2", "capital"),
        ],
        names=["Region", "Activity", "Factor of production"],
    )
    satellites = pd.MultiIndex.from_tuples(
        [("r1", "a1", "CO2"), ("r1", "a2", "CO2")],
        names=["Region", "Activity", "Satellite account"],
    )

    blocks = {
        "Z": pd.DataFrame(
            [
                [0.0, 0.0, 40.0, 10.0],
                [0.0, 0.0, 20.0, 30.0],
                [5.0, 7.0, 0.0, 0.0],
                [3.0, 9.0, 0.0, 0.0],
            ],
            index=productive,
            columns=productive,
        ),
        "Y": pd.DataFrame(
            [[0.0], [0.0], [50.0], [60.0]],
            index=productive,
            columns=final_demand,
        ),
        "V": pd.DataFrame(
            [
                [4.0, 0.0, 0.0, 0.0],
                [6.0, 0.0, 0.0, 0.0],
                [0.0, 5.0, 0.0, 0.0],
                [0.0, 8.0, 0.0, 0.0],
            ],
            index=factors,
            columns=productive,
        ),
        "E": pd.DataFrame(
            [
                [1.0, 0.0, 0.0, 0.0],
                [0.0, 2.0, 0.0, 0.0],
            ],
            index=satellites,
            columns=productive,
        ),
        "EY": pd.DataFrame([[0.5], [1.5]], index=satellites, columns=final_demand),
        "VY": pd.DataFrame([[0.0], [0.0], [0.0], [0.0]], index=factors, columns=final_demand),
    }
    indexes = {
        "r": {"main": ["r1"]},
        "n": {"main": ["hh"]},
        "a": {"main": ["a1", "a2"]},
        "c": {"main": ["c1", "c2"]},
        "s": {"main": ["a1", "a2", "c1", "c2"]},
        "f": {"main": ["taxes", "capital"]},
        "k": {"main": ["CO2"]},
    }
    units = {
        "Activity": pd.DataFrame({"unit": ["EUR", "EUR"]}, index=pd.Index(["a1", "a2"], name="Item")),
        "Commodity": pd.DataFrame({"unit": ["EUR", "EUR"]}, index=pd.Index(["c1", "c2"], name="Item")),
        "Factor of production": pd.DataFrame({"unit": ["EUR", "EUR"]}, index=pd.Index(["taxes", "capital"], name="Item")),
        "Satellite account": pd.DataFrame({"unit": ["ton"]}, index=pd.Index(["CO2"], name="Item")),
    }
    state = build_parser_state(
        table="SUT",
        matrices={"baseline": blocks},
        indexes=indexes,
        units=units,
        parser_name="tests",
        mode="flows",
        name="custom SUT",
    )
    state.metadata.extra["block_specs"] = sut_block_specs_for_matrix_layouts(
        {"V": ("Region", "Activity"), "E": ("Region", "Activity")}
    )
    return build_database_from_state(state, calc_all=False)


def _cvxlab_supports_csv_input_files():
    cvxlab = pytest.importorskip("cvxlab")
    from cvxlab.backend.model import Model

    return "input_data_files_type" in inspect.signature(Model.__init__).parameters

def test_get_corresponding_keys():

    # case 1, Sector
    item = _MASTER_INDEX["s"]

    keys,counter_item = get_corresponding_keys(item)

    assert counter_item == _MASTER_INDEX["s"]
    assert set(keys) == set([key for key in ADD_SECTOR_SHEETS.keys() if key not in ["of"]])

    # case 2, Activity
    item = _MASTER_INDEX["a"]

    keys,counter_item = get_corresponding_keys(item)

    assert counter_item == _MASTER_INDEX["c"]
    assert set(keys) == set([key for key in ADD_SECTOR_SHEETS.keys() if key not in ["sf","it"]])

    # case 3, Commodity
    item = _MASTER_INDEX["c"]

    keys,counter_item = get_corresponding_keys(item)

    assert counter_item == _MASTER_INDEX["a"]
    assert set(keys) == set([key for key in ADD_SECTOR_SHEETS.keys() if key not in ["sf","if"]])


def test_matrix_concat():

    # Test should be global for SUT and IOT

    set_1 = MatrixBuilder(
        "IOT",
        {
            _MASTER_INDEX.s : ["sector 1"],
            _MASTER_INDEX.k : ["CO2"],
            _MASTER_INDEX.n : ["FD"],
            _MASTER_INDEX.f : ["VA"],
            _MASTER_INDEX.r : ["reg 1","reg 2"]
        }
    )

    set_2 = MatrixBuilder(
        "IOT",
        {
            _MASTER_INDEX.s : ["sector 2"],
            _MASTER_INDEX.k : ["CO2"],
            _MASTER_INDEX.n : ["FD"],
            _MASTER_INDEX.f : ["VA"],
            _MASTER_INDEX.r : ["reg 1","reg 2"]
        }
    )

    data = dict(
        Y = set_1.Y,
        z = set_1.Z,
        e = set_1.E,
        v = set_1.V,
        EY = set_1.EY,
    )

    empty_matrices = dict(
        Y = set_2.Y,
        z = set_2.Z,
        e = set_2.E,
        v = set_2.V,
        EY = set_2.EY,
    )

    matrix_concat(data,empty_matrices)

    Y_index = pd.MultiIndex.from_product(
        [["reg 1","reg 2"],[_MASTER_INDEX["s"]],["sector 1","sector 2"]]
    ).sort_values()

    pdt.assert_index_equal(Y_index,data["Y"].index)
    pdt.assert_index_equal(Y_index,data["z"].index)
    pdt.assert_index_equal(Y_index,data["z"].columns)
    pdt.assert_index_equal(Y_index,data["v"].columns)
    pdt.assert_index_equal(Y_index,data["e"].columns)

def test_fill_matrix(CoreDataIOT):
    
    # Empty matrix
    

    empty_df = MatrixBuilder(
        "IOT",
        {
            _MASTER_INDEX.s : CoreDataIOT.get_index(_MASTER_INDEX.s) + ["new_sector"],
            _MASTER_INDEX.k : CoreDataIOT.get_index(_MASTER_INDEX.k),
            _MASTER_INDEX.n : CoreDataIOT.get_index(_MASTER_INDEX.n),
            _MASTER_INDEX.f : CoreDataIOT.get_index(_MASTER_INDEX.f),
            _MASTER_INDEX.r : CoreDataIOT.get_index(_MASTER_INDEX.r),
        }
    )

    user_df = getattr(CoreDataIOT,_ENUM.Z)
    output = fill_matrix(empty_df.Z,user_df)
    output.index.names = user_df.index.names
    output.columns.names = user_df.columns.names

    pdt.assert_frame_equal(
        user_df.sort_index(axis=0).sort_index(axis=1),
        output.drop(
            labels=["new_sector"],axis=0,level=-1
            ).drop(
                labels=["new_sector"],axis=1,level=-1
                ).sort_index(axis=0).sort_index(axis=1),
                
    )


def test_get_add_sectors_excel_supports_advanced_iot_workbook(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    new_sectors = ["New sector A", "New sector B"]
    path = tmp_path / "advanced_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=new_sectors,
        regions=regions,
        path=path,
        redefine_uncertainties=True,
    )

    workbook = pd.read_excel(path, sheet_name=None)
    item_sheet = ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET["IOT"]

    assert ADVANCED_ADD_SECTOR_MASTER_SHEET in workbook
    assert ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET in workbook
    assert ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET in workbook
    assert item_sheet in workbook
    assert "DB units" in workbook
    assert ADD_SECTOR_SPLIT_OUTPUT_SHEET in workbook
    assert ADD_SECTOR_SPLIT_TRADE_SHEET in workbook
    assert ADD_SECTOR_SPLIT_EXCLUSION_SHEET in workbook
    assert ADD_SECTOR_SPLIT_TOLERANCE_SHEET in workbook

    master = workbook[ADVANCED_ADD_SECTOR_MASTER_SHEET]
    assert master.columns.tolist() == list(
        ADVANCED_ADD_SECTOR_MASTER_SHEET_COLUMNS["IOT"].values()
    )
    assert master[_MASTER_INDEX["s"]].tolist() == [
        "New sector A",
        "New sector B",
        "New sector A",
        "New sector B",
    ]
    assert sorted(master["Inventory sheet"].tolist()) == [
        "INV_001",
        "INV_002",
        "INV_003",
        "INV_004",
    ]
    for sheet_name in master["Inventory sheet"].tolist():
        assert sheet_name in workbook


def test_get_add_sectors_excel_can_write_empty_workbook(tmp_path, CoreDataIOT):
    path = tmp_path / "empty_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(path=path)

    workbook = pd.read_excel(path, sheet_name=None)
    assert ADVANCED_ADD_SECTOR_MASTER_SHEET in workbook
    assert workbook[ADVANCED_ADD_SECTOR_MASTER_SHEET].empty


def test_get_add_sectors_excel_accepts_positional_path(tmp_path, CoreDataIOT):
    path = tmp_path / "positional_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(path)

    workbook = pd.read_excel(path, sheet_name=None)
    assert ADVANCED_ADD_SECTOR_MASTER_SHEET in workbook


def test_get_add_sectors_excel_refuses_to_overwrite_existing_workbook_by_default(
    tmp_path, CoreDataIOT
):
    path = tmp_path / "existing_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(path=path)

    with pytest.raises(WrongInput, match="already exists"):
        CoreDataIOT.get_add_sectors_excel(path=path)

    CoreDataIOT.get_add_sectors_excel(path=path, overwrite=True)

    workbook = pd.read_excel(path, sheet_name=None)
    assert ADVANCED_ADD_SECTOR_MASTER_SHEET in workbook


def test_get_add_sectors_excel_uses_all_regions_when_only_items_are_given(tmp_path, CoreDataIOT):
    path = tmp_path / "prefilled_all_regions_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector"],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET)
    assert sorted(master[_MASTER_INDEX["r"]].tolist()) == sorted(CoreDataIOT.get_index(_MASTER_INDEX["r"]))


def test_read_advanced_add_sector_workbook_roundtrip_parses_metadata(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
        redefine_uncertainties=True,
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame({"GLOBAL": regions, "EU": [regions[0], ""]}).to_excel(
            writer,
            sheet_name=ADVANCED_ADD_SECTOR_REGIONS_CLUSTERS_SHEET,
            index=False,
        )
        pd.DataFrame({"Cluster1": [CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]]}).to_excel(
            writer,
            sheet_name=ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET["IOT"],
            index=False,
        )
        pd.DataFrame(
            {
                "Inventory data categories": ["certain", "no info"],
                "New uncertainty values": [0.88, 0.22],
            }
        ).to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_UNCERTAINTIES_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 1,
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["s"]].iloc[0, 0],
                    "Input": "demo input",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": CoreDataIOT.get_index(_MASTER_INDEX["s"])[0],
                    "DB Region": regions[0],
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "note",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    workbook = read_advanced_add_sector_workbook(path, table="IOT")
    derived = derive_advanced_add_sector_sets(
        workbook,
        existing_sectors=CoreDataIOT.get_index(_MASTER_INDEX["s"]),
    )
    grouped = group_advanced_inventories_by_target(workbook)

    assert workbook.regions_clusters["GLOBAL"] == regions
    assert workbook.regions_clusters["EU"] == regions[:1]
    assert workbook.item_clusters["Cluster1"] == [CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]]
    assert workbook.uncertainty_values["certain"] == 0.88
    assert workbook.inventories_by_sheet["INV_001"].iloc[0]["DB Item"] == CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    assert derived["new_sectors"] == ["New sector A", "New sector B"]
    assert derived["to_split_sectors"] == []
    assert list(grouped) == ["New sector A", "New sector B"]
    assert "INV_001" in grouped["New sector A"]


def test_get_add_sectors_excel_supports_advanced_sut_workbook(tmp_path, CoreDataSUT):
    region = CoreDataSUT.get_index(_MASTER_INDEX["r"])[0]
    path = tmp_path / "advanced_add_sector_sut.xlsx"

    CoreDataSUT.get_add_sectors_excel(
        items=["New activity"],
        regions=[region],
        path=path,
        item=_MASTER_INDEX["a"],
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET)
    assert master[_MASTER_INDEX["a"]].tolist() == ["New activity"]
    assert master[_MASTER_INDEX["c"]].fillna("").tolist() == [""]


def test_get_add_sectors_excel_sut_defaults_to_activity_and_commodity(tmp_path, CoreDataSUT):
    region = CoreDataSUT.get_index(_MASTER_INDEX["r"])[0]
    path = tmp_path / "advanced_add_sector_sut_default.xlsx"

    CoreDataSUT.get_add_sectors_excel(
        items=["New item"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET)
    assert master[_MASTER_INDEX["a"]].tolist() == ["New item"]
    assert master[_MASTER_INDEX["c"]].tolist() == ["New item"]


def test_database_can_attach_advanced_add_sector_workbook_state(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
        redefine_uncertainties=True,
    )

    CoreDataIOT.read_add_sectors_excel(path)

    assert hasattr(CoreDataIOT, "add_sectors_workbook")
    assert CoreDataIOT.regions_clusters["GLOBAL"] == CoreDataIOT.get_index(_MASTER_INDEX["r"])
    assert CoreDataIOT.new_sectors == ["New sector A", "New sector B"]
    assert not hasattr(CoreDataIOT, "inventories")


def test_database_can_attach_inventories_via_read_add_sectors_excel(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot_with_inventories.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame(
            [
                {
                    "Quantity": 1,
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["s"]].iloc[0, 0],
                    "Input": "demo input",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": CoreDataIOT.get_index(_MASTER_INDEX["s"])[0],
                    "DB Region": regions[0],
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "note",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    ret = CoreDataIOT.read_add_sectors_excel(path, read_inventories=True)

    assert ret is None
    assert set(CoreDataIOT.inventories) == {"New sector A", "New sector B"}


def test_read_add_sectors_excel_can_merge_directory_inputs_with_renames(tmp_path, CoreDataIOT):
    directory = tmp_path / "multi_add_sectors"
    directory.mkdir()
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]

    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "a.xlsx",
        new_sector="New sector A",
        target_region=regions[0],
        inventory_sheet="Pippo",
        cluster_name="GLOBAL",
        cluster_members=[regions[0]],
    )
    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "b.xlsx",
        new_sector="New sector B",
        target_region=regions[1],
        inventory_sheet="Pippo",
        cluster_name="GLOBAL",
        cluster_members=[regions[1]],
    )
    (directory / "notes.txt").write_text("not a workbook")

    with pytest.warns(UserWarning, match=r"Skipping add-sectors file 'notes\.txt'"):
        CoreDataIOT.read_add_sectors_excel(directory, read_inventories=True)

    assert CoreDataIOT.new_sectors == ["New sector A", "New sector B"]
    assert set(CoreDataIOT.regions_clusters) == {"GLOBAL 1", "GLOBAL 2"}
    assert CoreDataIOT.regions_clusters["GLOBAL 1"] == [regions[0]]
    assert CoreDataIOT.regions_clusters["GLOBAL 2"] == [regions[1]]
    assert set(CoreDataIOT.add_sectors_master[_MASTER_INDEX["r"]]) == {"GLOBAL 1", "GLOBAL 2"}
    assert set(CoreDataIOT.inventories["New sector A"]) == {"Pippo 1"}
    assert set(CoreDataIOT.inventories["New sector B"]) == {"Pippo 2"}
    assert CoreDataIOT.inventories["New sector A"]["Pippo 1"].iloc[0]["DB Region"] == "GLOBAL 1"
    assert CoreDataIOT.inventories["New sector B"]["Pippo 2"].iloc[0]["DB Region"] == "GLOBAL 2"


def test_read_add_sectors_excel_keeps_one_identical_cluster_definition_from_directory(
    tmp_path, CoreDataIOT
):
    directory = tmp_path / "multi_add_sectors_same_cluster"
    directory.mkdir()
    all_regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])

    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "a.xlsx",
        new_sector="New sector A",
        target_region=all_regions[0],
        inventory_sheet="A inventory",
        cluster_name="GLOBAL",
        cluster_members=all_regions,
    )
    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "b.xlsx",
        new_sector="New sector B",
        target_region=all_regions[1],
        inventory_sheet="B inventory",
        cluster_name="GLOBAL",
        cluster_members=all_regions,
    )

    CoreDataIOT.read_add_sectors_excel(directory, read_inventories=True)

    assert list(CoreDataIOT.regions_clusters) == ["GLOBAL"]
    assert CoreDataIOT.regions_clusters["GLOBAL"] == all_regions


def test_read_add_sectors_excel_rejects_duplicate_targets_across_directory_workbooks(
    tmp_path, CoreDataIOT
):
    directory = tmp_path / "multi_add_sectors_duplicates"
    directory.mkdir()
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]

    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "a.xlsx",
        new_sector="New sector",
        target_region=regions[0],
        inventory_sheet="A inventory",
    )
    _configure_directory_add_sector_workbook(
        CoreDataIOT,
        directory / "b.xlsx",
        new_sector="New sector",
        target_region=regions[1],
        inventory_sheet="B inventory",
    )

    with pytest.raises(WrongInput, match="duplicate sectors across files"):
        CoreDataIOT.read_add_sectors_excel(directory, read_inventories=True)


def test_database_can_read_add_sectors_master_without_inventory_sheets(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot_master_only.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
    )

    workbook = load_workbook(path)
    for sheet_name in [name for name in workbook.sheetnames if name.startswith("INV_")]:
        workbook.remove(workbook[sheet_name])
    workbook.save(path)

    CoreDataIOT.read_add_sectors_excel(path, read_inventories=False)

    assert hasattr(CoreDataIOT, "add_sectors_workbook")
    assert CoreDataIOT.new_sectors == ["New sector A", "New sector B"]
    assert CoreDataIOT.add_sectors_workbook.inventories_by_sheet == {}
    assert not hasattr(CoreDataIOT, "inventories")


def test_get_inventory_sheets_creates_missing_inventory_tabs_from_master(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot_generate_inventories.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
    )

    workbook = load_workbook(path)
    for sheet_name in [name for name in workbook.sheetnames if name.startswith("INV_")]:
        workbook.remove(workbook[sheet_name])
    workbook.save(path)

    CoreDataIOT.read_add_sectors_excel(path)
    CoreDataIOT.get_inventory_sheets(path)

    workbook = load_workbook(path, data_only=False)
    assert {"INV_001", "INV_002", "DB units"} <= set(workbook.sheetnames)


def test_get_inventory_sheets_does_not_overwrite_existing_inventory_tabs_by_default(
    tmp_path, CoreDataIOT
):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:1]
    path = tmp_path / "advanced_add_sector_iot_preserve_inventories.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A"],
        regions=regions,
        path=path,
    )

    workbook = load_workbook(path)
    worksheet = workbook["INV_001"]
    worksheet["A2"] = "KEEP"
    workbook.save(path)

    CoreDataIOT.read_add_sectors_excel(path)
    CoreDataIOT.get_inventory_sheets(path)

    workbook = load_workbook(path, data_only=False)
    assert workbook["INV_001"]["A2"].value == "KEEP"


def test_read_add_sectors_excel_can_generate_inventory_templates(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:1]
    path = tmp_path / "advanced_add_sector_iot_get_inventories_flag.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A"],
        regions=regions,
        path=path,
    )

    workbook = load_workbook(path)
    for sheet_name in [name for name in workbook.sheetnames if name.startswith("INV_")]:
        workbook.remove(workbook[sheet_name])
    workbook.save(path)

    CoreDataIOT.read_add_sectors_excel(path, get_inventories=True)

    workbook = load_workbook(path, data_only=False)
    assert "INV_001" in workbook.sheetnames


def test_read_inventory_sheets_requires_missing_inventory_tabs(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:1]
    path = tmp_path / "advanced_add_sector_iot_missing_inventories.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A"],
        regions=regions,
        path=path,
    )

    workbook = load_workbook(path)
    for sheet_name in [name for name in workbook.sheetnames if name.startswith("INV_")]:
        workbook.remove(workbook[sheet_name])
    workbook.save(path)

    with pytest.raises(WrongExcelFormat, match="missing inventory sheets"):
        CoreDataIOT.read_inventory_sheets(path)


def test_database_can_group_inventories_from_advanced_workbook(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:2]
    path = tmp_path / "advanced_add_sector_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A", "New sector B"],
        regions=regions,
        path=path,
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame(
            [
                {
                    "Quantity": 1,
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["s"]].iloc[0, 0],
                    "Input": "demo input",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": CoreDataIOT.get_index(_MASTER_INDEX["s"])[0],
                    "DB Region": regions[0],
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "note",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    inventories = CoreDataIOT.read_inventory_sheets(path)
    assert "New sector A" in inventories
    assert inventories["New sector A"]["INV_001"].iloc[0]["Input"] == "demo input"


def test_read_inventory_sheets_evaluates_excel_formulas(tmp_path, CoreDataIOT):
    regions = CoreDataIOT.get_index(_MASTER_INDEX["r"])[:1]
    path = tmp_path / "advanced_add_sector_iot_formulas.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector A"],
        regions=regions,
        path=path,
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        pd.DataFrame({"Value": [3]}).to_excel(writer, sheet_name="Helper", index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": "='Helper'!A2*2",
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["s"]].iloc[0, 0],
                    "Input": "demo input",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": CoreDataIOT.get_index(_MASTER_INDEX["s"])[0],
                    "DB Region": regions[0],
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "note",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    inventories = CoreDataIOT.read_inventory_sheets(path)

    assert inventories["New sector A"]["INV_001"].iloc[0]["Quantity"] == 6


def test_add_sectors_advanced_engine_adds_iot_sector_from_workbook(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    existing_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    factor = CoreDataIOT.get_index(_MASTER_INDEX["f"])[0]
    satellite = CoreDataIOT.get_index(_MASTER_INDEX["k"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[existing_sector, "unit"]
    path = tmp_path / "advanced_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET)
    master = master.astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = existing_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.25,
                    "Unit": unit,
                    "Input": "input row",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": existing_sector,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 0.1,
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["f"]].loc[factor, "unit"],
                    "Input": "factor row",
                    "Item type": _MASTER_INDEX["f"],
                    "DB Item": factor,
                    "DB Region": "",
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 0.2,
                    "Unit": CoreDataIOT.units[_MASTER_INDEX["k"]].loc[satellite, "unit"],
                    "Input": "sat row",
                    "Item type": _MASTER_INDEX["k"],
                    "DB Item": satellite,
                    "DB Region": "",
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)

    assert "New sector" in new.get_index(_MASTER_INDEX["s"])
    assert new.units[_MASTER_INDEX["s"]].loc["New sector", "unit"] == unit
    assert new.z.loc[(region, _MASTER_INDEX["s"], existing_sector), (region, _MASTER_INDEX["s"], "New sector")] == pytest.approx(0.25)
    assert new.v.loc[factor, (region, _MASTER_INDEX["s"], "New sector")] == pytest.approx(0.1)
    assert new.e.loc[satellite, (region, _MASTER_INDEX["s"], "New sector")] == pytest.approx(0.2)
    assert hasattr(new, "uncertainty_matrix")
    assert new.uncertainty_matrix.loc[(region, _MASTER_INDEX["s"], existing_sector), (region, _MASTER_INDEX["s"], "New sector")] == 1


def test_add_sectors_logs_and_wraps_inventory_validation_errors(tmp_path, CoreDataIOT, caplog):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    existing_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[existing_sector, "unit"]
    path = tmp_path / "advanced_iot_invalid_units.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["New sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = existing_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.25,
                    "Unit": unit,
                    "Input": "bad sector row A",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": "Missing sector A",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 0.5,
                    "Unit": unit,
                    "Input": "bad sector row B",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": "Missing sector B",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    with caplog.at_level("ERROR", logger=database_module.logger.name):
        with pytest.raises(WrongInput) as msg:
            CoreDataIOT.add_sectors(io=path, inplace=False)

    message = str(msg.value)
    assert "Database.add_sectors failed during running add-sectors engine" in message
    assert "Issues found while validating inventory sheet INV_001" in message
    assert message.count("Database item could not be resolved") == 1
    assert "Excel row 2" in message
    assert "Excel row 3" in message
    assert "Missing sector A" in message
    assert "Missing sector B" in message
    assert any(
        "Database: add_sectors failed during running add-sectors engine" in record.message
        for record in caplog.records
    )


def test_add_sectors_supports_legacy_regional_extension_rows_iot(tmp_path):
    source_path = tmp_path / "legacy_regional_iot.xlsx"
    _write_legacy_regional_extension_iot(source_path)
    database = parse_from_excel(
        path=str(source_path),
        table="IOT",
        mode="flows",
        matrix_layouts={"E": "Region", "EY": "Region"},
        calc_all=False,
    )

    region = database.get_index(_MASTER_INDEX["r"])[0]
    other_region = database.get_index(_MASTER_INDEX["r"])[1]
    parent_sector = database.get_index(_MASTER_INDEX["s"])[0]
    factor = database.get_index(_MASTER_INDEX["f"])[0]
    satellite = database.get_index(_MASTER_INDEX["k"])[0]
    unit = database.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    path = tmp_path / "legacy_regional_add.xlsx"

    database.get_add_sectors_excel(items=["New sector"], regions=[region], path=path)
    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.1,
                    "Unit": database.units[_MASTER_INDEX["f"]].loc[factor, "unit"],
                    "Input": "factor row",
                    "Item type": _MASTER_INDEX["f"],
                    "DB Item": factor,
                    "DB Region": "",
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 7.0,
                    "Unit": database.units[_MASTER_INDEX["k"]].loc[satellite, "unit"],
                    "Input": "sat row",
                    "Item type": _MASTER_INDEX["k"],
                    "DB Item": satellite,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = database.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "New sector")
    parent_col = (region, _MASTER_INDEX["s"], parent_sector)

    assert new.E.index.names == ["Region", "Level", "Item"]
    assert new.v.loc[factor, new_col] == pytest.approx(0.1)
    assert new.e.loc[(region, "Satellite account", satellite), new_col] == pytest.approx(7.0)
    assert new.e.loc[(other_region, "Satellite account", satellite), new_col] == pytest.approx(
        database.e.loc[(other_region, "Satellite account", satellite), parent_col]
    )


def test_add_sectors_supports_factor_rows_with_region_sector_layout_iot(tmp_path):
    source_path = tmp_path / "legacy_layout_iot.xlsx"
    _write_legacy_layout_iot_with_regional_factors(source_path)
    database = parse_from_excel(
        path=str(source_path),
        table="IOT",
        mode="flows",
        matrix_layouts={"V": ("Region", "Sector"), "E": "Region", "EY": "Region"},
        calc_all=False,
    )

    region = database.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = database.get_index(_MASTER_INDEX["s"])[0]
    factor = "taxes"
    unit = database.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    path = tmp_path / "explicit_layout_add.xlsx"

    database.get_add_sectors_excel(items=["New sector"], regions=[region], path=path)
    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 11.0,
                    "Unit": database.units[_MASTER_INDEX["f"]].loc[factor, "unit"],
                    "Input": "factor row",
                    "Item type": _MASTER_INDEX["f"],
                    "DB Item": factor,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = database.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "New sector")
    parent_col = (region, _MASTER_INDEX["s"], parent_sector)

    expected = database.v.loc[(region, slice(None), "Factor of production", factor), parent_col]
    expected = expected / expected.sum() * 11.0
    allocated = new.v.loc[(region, slice(None), "Factor of production", factor), new_col]

    assert new.V.index.names == ["Region", "Sector", "Level", "Item"]
    assert set(new.V.index.tolist()) == set(database.V.index.tolist())
    assert allocated.sum() == pytest.approx(11.0)
    pdt.assert_series_equal(allocated.sort_index(), expected.sort_index(), check_names=False)


def test_add_sectors_supports_explicit_iot_axes_without_level_markers(tmp_path):
    source_path = tmp_path / "explicit_layout_iot_regional_fd.xlsx"
    _write_explicit_layout_iot_with_regional_final_demand(source_path)
    database = parse_from_excel(
        path=str(source_path),
        table="IOT",
        mode="flows",
        matrix_layouts={"V": ("Region", "Sector"), "E": "Region", "EY": "Region"},
        calc_all=False,
    )

    region = database.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = database.get_index(_MASTER_INDEX["s"])[0]
    factor = "taxes"
    unit = database.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    path = tmp_path / "explicit_axes_add.xlsx"

    database.get_add_sectors_excel(items=["New sector"], regions=[region], path=path)
    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 11.0,
                    "Unit": database.units[_MASTER_INDEX["f"]].loc[factor, "unit"],
                    "Input": "factor row",
                    "Item type": _MASTER_INDEX["f"],
                    "DB Item": factor,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = database.add_sectors(io=path, inplace=False)
    new_col = (region, "New sector")
    parent_col = (region, parent_sector)

    expected = database.v.loc[(region, slice(None), factor), parent_col]
    expected = expected / expected.sum() * 11.0
    allocated = new.v.loc[(region, slice(None), factor), new_col]

    assert new.z.index.names == ["Region", "Sector"]
    assert new.z.columns.names == ["Region", "Sector"]
    assert new.Y.index.names == ["Region", "Sector"]
    assert new.Y.columns.names == ["Region", "Consumption category"]
    assert new.v.index.names == ["Region", "Sector", "Factor of production"]
    assert new.v.columns.names == ["Region", "Sector"]
    assert new.e.index.names == ["Region", "Satellite account"]
    assert new.e.columns.names == ["Region", "Sector"]
    assert ("r1", "New sector") in new.z.columns.tolist()
    assert allocated.sum() == pytest.approx(11.0)
    pdt.assert_series_equal(allocated.sort_index(), expected.sort_index(), check_names=False)


def test_add_sectors_returns_none_when_inplace_true(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    existing_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[existing_sector, "unit"]
    path = tmp_path / "inplace_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Inplace sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = existing_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)

    result = CoreDataIOT.add_sectors(io=path, inplace=True)

    assert result is None
    assert "Inplace sector" in CoreDataIOT.get_index(_MASTER_INDEX["s"])


def test_add_sectors_advanced_engine_adds_sut_activity_and_commodity(tmp_path, CoreDataSUT):
    region = CoreDataSUT.get_index(_MASTER_INDEX["r"])[0]
    existing_activity = CoreDataSUT.get_index(_MASTER_INDEX["a"])[0]
    existing_commodity = CoreDataSUT.get_index(_MASTER_INDEX["c"])[0]
    unit = CoreDataSUT.units[_MASTER_INDEX["a"]].loc[existing_activity, "unit"]
    path = tmp_path / "advanced_sut.xlsx"

    CoreDataSUT.get_add_sectors_excel(
        items=["New activity"],
        regions=[region],
        path=path,
        item=_MASTER_INDEX["a"],
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET)
    master = master.astype(object)
    master.loc[0, "Commodity"] = "New commodity"
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Activity"] = existing_activity
    master.loc[0, "Market share"] = 1.0
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.4,
                    "Unit": CoreDataSUT.units[_MASTER_INDEX["c"]].loc[existing_commodity, "unit"],
                    "Input": "commodity row",
                    "Item type": _MASTER_INDEX["c"],
                    "DB Item": existing_commodity,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataSUT.add_sectors(io=path, inplace=False)

    assert "New activity" in new.get_index(_MASTER_INDEX["a"])
    assert "New commodity" in new.get_index(_MASTER_INDEX["c"])
    assert new.units[_MASTER_INDEX["a"]].loc["New activity", "unit"] == unit
    assert new.units[_MASTER_INDEX["c"]].loc["New commodity", "unit"] == unit
    assert new.u.loc[(region, _MASTER_INDEX["c"], existing_commodity), (region, _MASTER_INDEX["a"], "New activity")] == pytest.approx(0.4)
    assert new.s.loc[(region, _MASTER_INDEX["a"], "New activity"), (region, _MASTER_INDEX["c"], "New commodity")] == pytest.approx(1.0)


def test_add_sectors_reports_inventory_validation_errors_for_sut(tmp_path, CoreDataSUT):
    region = CoreDataSUT.get_index(_MASTER_INDEX["r"])[0]
    existing_activity = CoreDataSUT.get_index(_MASTER_INDEX["a"])[0]
    unit = CoreDataSUT.units[_MASTER_INDEX["a"]].loc[existing_activity, "unit"]
    path = tmp_path / "advanced_sut_invalid_inventory.xlsx"

    CoreDataSUT.get_add_sectors_excel(
        items=["New activity"],
        regions=[region],
        path=path,
        item=_MASTER_INDEX["a"],
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Commodity"] = "New commodity"
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Activity"] = existing_activity
    master.loc[0, "Market share"] = 1.0
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.4,
                    "Unit": unit,
                    "Input": "bad commodity row",
                    "Item type": _MASTER_INDEX["c"],
                    "DB Item": "Missing commodity",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    with pytest.raises(WrongInput) as msg:
        CoreDataSUT.add_sectors(io=path, inplace=False)

    message = str(msg.value)
    assert "Database.add_sectors failed during running add-sectors engine" in message
    assert "Issues found while validating inventory sheet INV_001" in message
    assert "Database item could not be resolved" in message
    assert "Missing commodity" in message


def test_add_sectors_advanced_engine_supports_custom_sut_factor_and_extension_rows(tmp_path):
    database = _build_custom_sut_database()
    region = database.get_index(_MASTER_INDEX["r"])[0]
    existing_activity = database.get_index(_MASTER_INDEX["a"])[0]
    existing_commodity = database.get_index(_MASTER_INDEX["c"])[0]
    unit = database.units[_MASTER_INDEX["a"]].loc[existing_activity, "unit"]
    path = tmp_path / "advanced_custom_sut.xlsx"

    database.get_add_sectors_excel(
        items=["New activity"],
        regions=[region],
        path=path,
        item=_MASTER_INDEX["a"],
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Commodity"] = "New commodity"
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Activity"] = existing_activity
    master.loc[0, "Market share"] = 1.0
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.4,
                    "Unit": database.units[_MASTER_INDEX["c"]].loc[existing_commodity, "unit"],
                    "Input": "commodity row",
                    "Item type": _MASTER_INDEX["c"],
                    "DB Item": existing_commodity,
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 0.7,
                    "Unit": database.units[_MASTER_INDEX["f"]].loc["taxes", "unit"],
                    "Input": "factor row",
                    "Item type": _MASTER_INDEX["f"],
                    "DB Item": "taxes",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
                {
                    "Quantity": 0.2,
                    "Unit": database.units[_MASTER_INDEX["k"]].loc["CO2", "unit"],
                    "Input": "sat row",
                    "Item type": _MASTER_INDEX["k"],
                    "DB Item": "CO2",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                },
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = database.add_sectors(io=path, inplace=False)

    assert new.V.index.names == ["Region", "Activity", "Factor of production"]
    assert new.E.index.names == ["Region", "Activity", "Satellite account"]
    assert new.u.loc[(region, _MASTER_INDEX["c"], existing_commodity), (region, _MASTER_INDEX["a"], "New activity")] == pytest.approx(0.4)
    assert new.v.loc[(region, existing_activity, "taxes"), (region, _MASTER_INDEX["a"], "New activity")] == pytest.approx(0.7)
    assert new.e.loc[(region, existing_activity, "CO2"), (region, _MASTER_INDEX["a"], "New activity")] == pytest.approx(0.2)


def test_add_sectors_advanced_engine_honors_leave_empty_iot(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    existing_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[existing_sector, "unit"]
    path = tmp_path / "advanced_iot_leave_empty.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Empty sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Leave empty"] = True
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "Empty sector")
    new_row = (region, _MASTER_INDEX["s"], "Empty sector")

    assert new.z.loc[:, new_col].sum() == pytest.approx(0.0)
    assert new.z.loc[new_row, :].sum() == pytest.approx(0.0)


def test_add_sectors_advanced_engine_supports_percentage_updates_iot(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    parent_column = CoreDataIOT.z.loc[:, (region, _MASTER_INDEX["s"], parent_sector)]
    nonzero_rows = parent_column[parent_column != 0]
    target_row = nonzero_rows.index[0]
    untouched_row = nonzero_rows.index[1]
    path = tmp_path / "advanced_iot_percentage.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Scaled sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 0.5,
                    "Unit": unit,
                    "Input": "scaled row",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": target_row[2],
                    "DB Region": target_row[0],
                    "Change type": "Percentage",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "Scaled sector")

    assert new.z.loc[target_row, new_col] == pytest.approx(parent_column.loc[target_row] * 1.5)
    assert new.z.loc[untouched_row, new_col] == pytest.approx(parent_column.loc[untouched_row])


def test_add_sectors_advanced_engine_supports_region_cluster_updates_iot(tmp_path, CoreDataIOT):
    target_region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    sector_candidates = CoreDataIOT.get_index(_MASTER_INDEX["s"])
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    chosen_sector = None
    for sector in sector_candidates:
        series = CoreDataIOT.Z.loc[
            (slice(None), _MASTER_INDEX["s"], sector),
            (target_region, _MASTER_INDEX["s"], parent_sector),
        ]
        if (series > 0).sum() >= 2:
            chosen_sector = sector
            break
    assert chosen_sector is not None
    path = tmp_path / "advanced_iot_cluster.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Clustered sector"],
        regions=[target_region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 10.0,
                    "Unit": unit,
                    "Input": "cluster row",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": chosen_sector,
                    "DB Region": "GLOBAL",
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)
    new_col = (target_region, _MASTER_INDEX["s"], "Clustered sector")
    allocated = new.z.loc[(slice(None), _MASTER_INDEX["s"], chosen_sector), new_col]
    assert allocated.sum() == pytest.approx(10.0)
    assert (allocated > 0).sum() >= 2


def test_add_sectors_advanced_engine_converts_satellite_units_iot(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    sector_unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    satellite = CoreDataIOT.units[_MASTER_INDEX["k"]].query("unit == 'kg'").index[0]
    path = tmp_path / "advanced_iot_units.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Converted sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = sector_unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(
            [
                {
                    "Quantity": 1000.0,
                    "Unit": "g",
                    "Input": "converted energy",
                    "Item type": _MASTER_INDEX["k"],
                    "DB Item": satellite,
                    "DB Region": "",
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "Converted sector")

    assert new.e.loc[satellite, new_col] == pytest.approx(1.0)


def test_add_sectors_advanced_engine_supports_item_cluster_updates_iot(tmp_path, CoreDataIOT):
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    sector_unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]
    positive_inputs = []
    for sector in CoreDataIOT.get_index(_MASTER_INDEX["s"]):
        total = CoreDataIOT.Z.loc[
            (region, _MASTER_INDEX["s"], sector),
            (region, slice(None), slice(None)),
        ].sum()
        if total > 0:
            positive_inputs.append(sector)
        if len(positive_inputs) == 2:
            break

    assert len(positive_inputs) == 2
    path = tmp_path / "advanced_iot_item_cluster.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Cluster sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = sector_unit
    master.loc[0, "Parent Sector"] = parent_sector
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame({"Bundle": positive_inputs}).to_excel(
            writer,
            sheet_name=ADVANCED_ADD_SECTOR_ITEMS_CLUSTERS_SHEET["IOT"],
            index=False,
        )
        pd.DataFrame(
            [
                {
                    "Quantity": 10.0,
                    "Unit": sector_unit,
                    "Input": "clustered input",
                    "Item type": _MASTER_INDEX["s"],
                    "DB Item": "Bundle",
                    "DB Region": region,
                    "Change type": "Update",
                    "Source": "test",
                    "Notes": "",
                }
            ]
        ).to_excel(writer, sheet_name="INV_001", index=False)

    new = CoreDataIOT.add_sectors(io=path, inplace=False)
    new_col = (region, _MASTER_INDEX["s"], "Cluster sector")
    allocated = new.z.loc[(region, _MASTER_INDEX["s"], positive_inputs), new_col]

    assert allocated.sum() == pytest.approx(10.0)
    assert (allocated > 0).sum() == 2


def test_add_sectors_advanced_engine_supports_sut_final_demand_from_master(tmp_path, CoreDataSUT):
    region = CoreDataSUT.get_index(_MASTER_INDEX["r"])[0]
    new_item = "Demanded item"
    activity_unit = CoreDataSUT.units[_MASTER_INDEX["a"]].iloc[0, 0]
    final_demand = CoreDataSUT.get_index(_MASTER_INDEX["n"])[0]
    path = tmp_path / "advanced_sut_final_demand.xlsx"

    CoreDataSUT.get_add_sectors_excel(
        items=[new_item],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = activity_unit
    master.loc[0, "Market share"] = 0.75
    master.loc[0, "Final consumption"] = 2.5
    master.loc[0, _MASTER_INDEX["n"]] = final_demand
    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)

    new = CoreDataSUT.add_sectors(io=path, inplace=False)
    new_supply_row = (region, _MASTER_INDEX["a"], new_item)
    new_use_row = (region, _MASTER_INDEX["c"], new_item)
    final_demand_col = (region, _MASTER_INDEX["n"], final_demand)

    assert new.s.loc[new_supply_row, new_use_row] == pytest.approx(0.75)
    assert new.Y.loc[new_use_row, final_demand_col] == pytest.approx(2.5)


def test_read_add_sectors_excel_attaches_split_metadata(tmp_path, CoreDataIOT):
    path = tmp_path / "split_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    split_setup = _configure_split_workbook(CoreDataIOT, path)

    CoreDataIOT.read_add_sectors_excel(path)

    assert CoreDataIOT.to_split_sectors == [split_setup["new_sector"]]
    assert ADD_SECTOR_SPLIT_OUTPUT_SHEET in CoreDataIOT.split_info
    assert ADD_SECTOR_SPLIT_TRADE_SHEET in CoreDataIOT.split_info
    assert (
        CoreDataIOT.split_info[ADD_SECTOR_SPLIT_OUTPUT_SHEET].iloc[0][
            ADD_SECTOR_SPLIT_OUTPUT_COLUMNS["sector"]
        ]
        == split_setup["new_sector"]
    )


def test_add_sectors_split_fails_before_engine_on_invalid_total_output_sheet(
    tmp_path, CoreDataIOT, monkeypatch
):
    path = tmp_path / "invalid_split_iot.xlsx"
    region = CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]
    parent_sector = CoreDataIOT.get_index(_MASTER_INDEX["s"])[0]
    unit = CoreDataIOT.units[_MASTER_INDEX["s"]].loc[parent_sector, "unit"]

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[region],
        path=path,
    )

    master = pd.read_excel(path, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET).astype(object)
    master.loc[0, "Unit"] = unit
    master.loc[0, "Parent Sector"] = parent_sector
    master.loc[0, "Add or Split"] = "Split"
    split_trades = pd.DataFrame(
        [
            {
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["sector_from"]: "Split sector",
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_from"]: region,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_to"]: region,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["quantity"]: 0.1,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["unit"]: unit,
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["source"]: "test",
                ADD_SECTOR_SPLIT_TRADE_COLUMNS["notes"]: "",
            }
        ]
    )

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        master.to_excel(writer, sheet_name=ADVANCED_ADD_SECTOR_MASTER_SHEET, index=False)
        pd.DataFrame(columns=list(ADD_SECTOR_SPLIT_OUTPUT_COLUMNS.values())).to_excel(
            writer, sheet_name=ADD_SECTOR_SPLIT_OUTPUT_SHEET, index=False
        )
        split_trades.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET, index=False)

    def _should_not_run(*args, **kwargs):
        raise AssertionError("run_add_sector_engine should not be called for invalid split support")

    monkeypatch.setattr(database_module, "run_add_sector_engine", _should_not_run)

    with pytest.raises(WrongInput, match="missing total output rows"):
        CoreDataIOT.add_sectors(
            io=path,
            inplace=False,
            split=True,
            cvxlab_path=tmp_path,
            only_input_data_gen=True,
        )


def test_add_sectors_split_can_generate_cvxlab_input_data(tmp_path, CoreDataIOT):
    pytest.importorskip("cvxlab")
    path = tmp_path / "split_iot.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    split_setup = _configure_split_workbook(CoreDataIOT, path)

    new = CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        only_input_data_gen=True,
    )

    model_dir = tmp_path / CVXLAB_SPLIT_MODEL_NAME
    input_file = model_dir / "input_data" / "input_data.xlsx"
    assert set(new.matrices) == {"baseline"}
    assert split_setup["new_sector"] in new.get_index(_MASTER_INDEX["s"])
    assert model_dir.exists()
    assert input_file.exists()

    workbook = pd.ExcelFile(input_file)
    assert {"V0", "Vold", "Z0", "Zold", "Yold", "Trade", "Trade_selector", "tol"}.issubset(
        set(workbook.sheet_names)
    )
    vold = pd.read_excel(input_file, sheet_name="Vold")
    assert {"factor_Name", "region_to_Name", "sector_to_Name", "values"}.issubset(vold.columns)


def test_add_sectors_split_keeps_domestic_trade_selector_rows_zero(tmp_path, CoreDataIOT):
    pytest.importorskip("cvxlab")
    path = tmp_path / "split_iot_domestic_trade.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    split_setup = _configure_split_workbook(CoreDataIOT, path)

    split_trades = pd.read_excel(path, sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET)
    split_trades[ADD_SECTOR_SPLIT_TRADE_COLUMNS["region_to"]] = split_setup["region"]

    with pd.ExcelWriter(path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        split_trades.to_excel(writer, sheet_name=ADD_SECTOR_SPLIT_TRADE_SHEET, index=False)

    CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        only_input_data_gen=True,
    )

    input_file = tmp_path / CVXLAB_SPLIT_MODEL_NAME / "input_data" / "input_data.xlsx"
    trade_selector = pd.read_excel(input_file, sheet_name="Trade_selector")
    same_region = trade_selector[
        trade_selector["region_from_Name"] == trade_selector["region_to_Name"]
    ]

    assert not same_region.empty
    assert same_region["values"].eq(0).all()


def test_add_sectors_split_zeroes_sub_threshold_input_values(tmp_path, CoreDataIOT):
    pytest.importorskip("cvxlab")
    path = tmp_path / "split_iot_tiny_values.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    _configure_split_workbook(CoreDataIOT, path, quantity=1e-8)

    CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        only_input_data_gen=True,
    )

    input_file = tmp_path / CVXLAB_SPLIT_MODEL_NAME / "input_data" / "input_data.xlsx"
    workbook = pd.ExcelFile(input_file)

    for sheet_name in workbook.sheet_names:
        if sheet_name in {"Trade_selector", "tol"}:
            continue
        df = pd.read_excel(input_file, sheet_name=sheet_name)
        if "values" not in df.columns:
            continue
        small_positive_values = df.loc[(df["values"] > 0) & (df["values"] < 1e-6), "values"]
        assert small_positive_values.empty, f"{sheet_name} still contains positive values below 1e-6"


def test_add_sectors_split_old_tables_include_non_split_added_sectors(tmp_path):
    pytest.importorskip("cvxlab")
    database = load_test("IOT")
    workbook = Path("mario/test/supporting_files/add_sector_iot_inventories_filled.xlsx")

    new = database.add_sectors(
        io=workbook,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        only_input_data_gen=True,
    )

    input_file = tmp_path / CVXLAB_SPLIT_MODEL_NAME / "input_data" / "input_data.xlsx"
    zold = pd.read_excel(input_file, sheet_name="Zold")
    yold = pd.read_excel(input_file, sheet_name="Yold")
    vold = pd.read_excel(input_file, sheet_name="Vold")

    assert set(new.matrices) == {"baseline"}
    assert not zold["values"].isna().any()
    assert not yold["values"].isna().any()
    assert not vold["values"].isna().any()


def test_add_sectors_split_can_generate_cvxlab_csv_input_data(tmp_path, CoreDataIOT):
    if not _cvxlab_supports_csv_input_files():
        pytest.skip("Installed CVXLab build does not support csv input files.")

    path = tmp_path / "split_iot_csv.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    _configure_split_workbook(CoreDataIOT, path)

    new = CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        only_input_data_gen=True,
        input_data_files_type="csv",
    )

    model_dir = tmp_path / CVXLAB_SPLIT_MODEL_NAME / "input_data"
    assert set(new.matrices) == {"baseline"}
    assert (model_dir / "Trade.csv").exists()
    assert (model_dir / "tol.csv").exists()


def test_add_sectors_split_rejects_custom_extension_layouts(tmp_path):
    source_path = tmp_path / "legacy_regional_iot.xlsx"
    _write_legacy_regional_extension_iot(source_path)
    database = parse_from_excel(
        path=str(source_path),
        table="IOT",
        mode="flows",
        matrix_layouts={"E": "Region", "EY": "Region"},
        calc_all=False,
    )

    path = tmp_path / "split_custom_e.xlsx"
    database.get_add_sectors_excel(items=["Split sector"], regions=[database.get_index(_MASTER_INDEX["r"])[0]], path=path)
    _configure_split_workbook(database, path)

    with pytest.raises(NotImplementable, match="split=True currently supports only classic CVXLab row layouts"):
        database.add_sectors(
            io=path,
            inplace=False,
            split=True,
            cvxlab_path=tmp_path,
            only_input_data_gen=True,
        )


def test_add_sectors_split_rejects_custom_factor_layouts(tmp_path):
    source_path = tmp_path / "legacy_layout_iot.xlsx"
    _write_legacy_layout_iot_with_regional_factors(source_path)
    database = parse_from_excel(
        path=str(source_path),
        table="IOT",
        mode="flows",
        matrix_layouts={"V": ("Region", "Sector"), "E": "Region", "EY": "Region"},
        calc_all=False,
    )

    path = tmp_path / "split_custom_v.xlsx"
    database.get_add_sectors_excel(items=["Split sector"], regions=[database.get_index(_MASTER_INDEX["r"])[0]], path=path)
    _configure_split_workbook(database, path)

    with pytest.raises(NotImplementable, match="split=True currently supports only classic CVXLab row layouts"):
        database.add_sectors(
            io=path,
            inplace=False,
            split=True,
            cvxlab_path=tmp_path,
            only_input_data_gen=True,
        )


def test_split_bridge_uses_historical_coordinate_loading_flow(tmp_path):
    class _FakeModel:
        def __init__(self):
            self.called = []

        def initialize_model_environment(self):
            raise AssertionError("MARIO should not delegate split setup to initialize_model_environment().")

        def load_model_coordinates(self):
            self.called.append("load_model_coordinates")

        def initialize_blank_data_structure(self):
            self.called.append("initialize_blank_data_structure")

    model = _FakeModel()

    _load_model_coordinates_and_initialize_database(model, dest_dir=tmp_path)

    assert model.called == [
        "load_model_coordinates",
        "initialize_blank_data_structure",
    ]


def test_split_bridge_copies_model_settings_verbatim(tmp_path):
    source = Path("mario/ops/cvxlab_models/Split_sectors/model_settings.xlsx")
    dest = bridge_module._prepare_split_model_directory(
        main_dir_path=tmp_path,
        model_dir_name=CVXLAB_SPLIT_MODEL_NAME,
    )
    copied = dest / "model_settings.xlsx"
    if getattr(bridge_module.cl.Defaults.Labels, "NONNEG_KEY", None) is not None:
        assert hashlib.md5(source.read_bytes()).hexdigest() == hashlib.md5(copied.read_bytes()).hexdigest()
    else:
        source_df = pd.read_excel(source, sheet_name="structure_variables")
        copied_df = pd.read_excel(copied, sheet_name="structure_variables")
        assert "nonneg" in source_df.columns
        assert "nonneg" not in copied_df.columns


def test_split_bridge_fills_blank_problem_cells_for_newer_cvxlab_builds(tmp_path, monkeypatch):
    source = Path("mario/ops/cvxlab_models/Split_sectors/model_settings.xlsx")
    copied = tmp_path / "model_settings.xlsx"
    copied.write_bytes(source.read_bytes())

    class _FakeModel:
        refresh_database_and_initialize_problem = object()

    class _FakeLabels:
        NONNEG_KEY = "nonneg"

    class _FakeDefaults:
        Labels = _FakeLabels

    class _FakeCL:
        Model = _FakeModel
        Defaults = _FakeDefaults

    monkeypatch.setattr(bridge_module, "cl", _FakeCL)

    bridge_module._maybe_rewrite_incompatible_model_settings(copied)

    problem = pd.read_excel(copied, sheet_name="problem", keep_default_na=False)
    assert all(isinstance(value, str) for value in problem["objective"])
    assert all(isinstance(value, str) for value in problem["expressions"])


def test_split_bridge_sanitizes_loaded_symbolic_problem_entries():
    payload = {
        "objective": ["Minimize(x)", float("nan"), "", "   "],
        "expressions": ["x >= 0", None, " ", float("nan")],
        "nested": {
            "objective": [None, "Minimize(y)"],
            "expressions": ["y >= 0", ""],
        },
    }

    bridge_module._sanitize_symbolic_problem_structure(payload)

    assert payload["objective"] == ["Minimize(x)"]
    assert payload["expressions"] == ["x >= 0"]
    assert payload["nested"]["objective"] == ["Minimize(y)"]
    assert payload["nested"]["expressions"] == ["y >= 0"]


def test_split_bridge_keeps_description_text_raw_while_parsing_structured_fields():
    data = pd.DataFrame(
        [
            {
                "table_key": "XT",
                "description": "New total production (r_to,s_to)",
                "type": "endogenous",
                "integer": None,
                "coordinates": "region_to, sector_to",
                "variables_info": "XT_n",
            }
        ]
    )

    result = bridge_module._pivot_cvxlab_settings_dataframe(data, primary_key="table_key")

    assert result["XT"]["description"] == "New total production (r_to,s_to)"
    assert result["XT"]["coordinates"] == ["region_to", "sector_to"]
    assert result["XT"]["variables_info"] == "XT_n"


def test_split_bridge_normalizes_relative_cvxlab_root(monkeypatch):
    captured = {}

    class _FakeModel:
        def __init__(self, **kwargs):
            captured.update(kwargs)

    class _FakeCL:
        Model = _FakeModel

    monkeypatch.setattr(bridge_module, "cl", _FakeCL)

    _build_cvxlab_model(
        main_dir_path="relative_split_root",
        model_dir_name="model_dir",
        input_data_files_type="xlsx",
    )

    assert captured["main_dir_path"] == str((Path.cwd() / "relative_split_root").resolve())


def test_split_bridge_rejects_cvxlab_tables_missing_required_coordinates():
    base = pd.DataFrame(
        {
            "id": [1, 2],
            "region_to_Name": ["Reg1", "Reg2"],
            "sector_to_Name": ["New industry", "New industry"],
            "values": [0.0, 0.0],
        }
    )
    updates = pd.DataFrame(
        {
            "factors_Name": ["Taxes", "Wages", "Capital", "Taxes", "Wages", "Capital"],
            "region_to_Name": ["Reg1", "Reg1", "Reg1", "Reg2", "Reg2", "Reg2"],
            "sector_to_Name": ["New industry"] * 6,
            "values": [1.0, 2.0, 3.0, 4.0, 5.0, 6.0],
        }
    )

    with pytest.raises(NotImplementable, match="input table 'VA'.*factors_Name"):
        bridge_module._merge_cvxlab_input_table(base, updates, table_name="VA")


def test_split_bridge_picks_conic_solver_by_default(monkeypatch):
    cp = pytest.importorskip("cvxpy")

    monkeypatch.setattr(cp, "installed_solvers", lambda: ["SCIPY", "SCS"])

    assert bridge_module._resolve_split_solver(None) == "SCS"
    assert bridge_module._resolve_split_solver("GUROBI") == "GUROBI"


def test_split_bridge_accepts_multi_subproblem_status_dict():
    status = {
        "Sub-problem [0]": "optimal",
        "Sub-problem [1]": "optimal_inaccurate",
    }

    assert bridge_module._cvxlab_problem_solved_optimally(status)


def test_split_bridge_rejects_non_optimal_multi_subproblem_status_dict():
    status = {
        "Sub-problem [0]": "optimal",
        "Sub-problem [1]": "infeasible",
    }

    assert not bridge_module._cvxlab_problem_solved_optimally(status)


def test_add_sectors_split_can_attach_mocked_cvxlab_results(tmp_path, CoreDataIOT, monkeypatch):
    path = tmp_path / "split_iot_optimized.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    _configure_split_workbook(CoreDataIOT, path)

    def _fake_optimize(instance, **kwargs):
        split_baseline = instance.matrices["split_baseline"]
        return {
            _ENUM["Z"]: split_baseline[_ENUM["Z"]].copy(),
            _ENUM["Y"]: split_baseline[_ENUM["Y"]].copy(),
            _ENUM["V"]: split_baseline[_ENUM["V"]].copy(),
        }

    monkeypatch.setattr(database_module, "optimize_split_in_cvxlab", _fake_optimize)

    new = CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
    )

    assert set(new.matrices) == {"baseline"}
    assert _ENUM["X"] in new.matrices["baseline"]


def test_add_sectors_split_can_keep_all_intermediate_scenarios(tmp_path, CoreDataIOT, monkeypatch):
    path = tmp_path / "split_iot_keep_steps.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    _configure_split_workbook(CoreDataIOT, path)

    def _fake_optimize(instance, **kwargs):
        split_baseline = instance.matrices["split_baseline"]
        return {
            _ENUM["Z"]: split_baseline[_ENUM["Z"]].copy(),
            _ENUM["Y"]: split_baseline[_ENUM["Y"]].copy(),
            _ENUM["V"]: split_baseline[_ENUM["V"]].copy(),
        }

    monkeypatch.setattr(database_module, "optimize_split_in_cvxlab", _fake_optimize)

    new = CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        keep_all_split_steps=True,
        cvxlab_path=tmp_path,
    )

    assert set(new.matrices) == {"baseline", "original", "split_baseline", "split_cvxlab"}
    assert _ENUM["X"] in new.matrices["split_cvxlab"]


def test_add_sectors_split_can_rename_parent_sector(tmp_path, CoreDataIOT, monkeypatch):
    path = tmp_path / "split_iot_parent_name.xlsx"

    CoreDataIOT.get_add_sectors_excel(
        items=["Split sector"],
        regions=[CoreDataIOT.get_index(_MASTER_INDEX["r"])[0]],
        path=path,
    )
    split_setup = _configure_split_workbook(CoreDataIOT, path)
    parent_name = f"Other {split_setup['parent_sector']}"

    def _fake_optimize(instance, **kwargs):
        split_baseline = instance.matrices["split_baseline"]
        return {
            _ENUM["Z"]: split_baseline[_ENUM["Z"]].copy(),
            _ENUM["Y"]: split_baseline[_ENUM["Y"]].copy(),
            _ENUM["V"]: split_baseline[_ENUM["V"]].copy(),
        }

    monkeypatch.setattr(database_module, "optimize_split_in_cvxlab", _fake_optimize)

    new = CoreDataIOT.add_sectors(
        io=path,
        inplace=False,
        split=True,
        cvxlab_path=tmp_path,
        parent_name=parent_name,
    )

    assert parent_name in new.get_index(_MASTER_INDEX["s"])
    assert split_setup["parent_sector"] not in new.get_index(_MASTER_INDEX["s"])
    assert parent_name in new.matrices["baseline"][_ENUM["Z"]].index.get_level_values(2)
    assert parent_name in new.matrices["baseline"][_ENUM["Z"]].columns.get_level_values(2)
    assert parent_name in new.units[_MASTER_INDEX["s"]].index
