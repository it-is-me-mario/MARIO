from __future__ import annotations

from html import escape
from pathlib import Path

from openpyxl import load_workbook


DOC_ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = DOC_ROOT / "source" / "_static" / "data" / "supporting_files" / "Terminology.xlsx"
OUTPUT_DIR = DOC_ROOT / "source" / "concepts" / "_generated"


def normalize_rows(sheet_name: str):
    workbook = load_workbook(WORKBOOK, data_only=True)
    worksheet = workbook[sheet_name]
    rows = list(worksheet.iter_rows(values_only=True))
    headers = [str(value).strip() for value in rows[0] if value not in (None, "")]
    width = len(headers)
    body = []
    for row in rows[1:]:
        values = list(row[:width])
        if all(value in (None, "") for value in values):
            continue
        body.append(["" if value is None else str(value).strip() for value in values])
    return headers, body


def render_table(table_id: str, headers: list[str], rows: list[list[str]]) -> str:
    html = []
    html.append(f'<div class="terminology-table-wrapper">')
    html.append(f'  <table id="{table_id}" class="terminology-table">')
    html.append("    <thead>")
    html.append("      <tr>")
    for header in headers:
        html.append(f"        <th>{escape(header)}</th>")
    html.append("      </tr>")
    html.append('      <tr class="terminology-column-filters">')
    for index, header in enumerate(headers):
        html.append("        <th>")
        html.append(
            f'          <input class="terminology-column-filter" data-col-index="{index}" '
            f'type="search" placeholder="Filter {escape(header)}..." />'
        )
        html.append("        </th>")
    html.append("      </tr>")
    html.append("    </thead>")
    html.append("    <tbody>")
    for row in rows:
        search_blob = " ".join(row).lower()
        html.append(f'      <tr data-search="{escape(search_blob)}">')
        for index, cell in enumerate(row):
            if headers and headers[0] == "Matrix" and index == 0:
                html.append(f"        <td><strong>{escape(cell)}</strong></td>")
            else:
                html.append(f"        <td>{escape(cell)}</td>")
        html.append("      </tr>")
    html.append("    </tbody>")
    html.append("  </table>")
    html.append("</div>")
    return "\n".join(html) + "\n"


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    indices_headers, indices_rows = normalize_rows("Indices")
    matrices_headers, matrices_rows = normalize_rows("Matrices")

    (OUTPUT_DIR / "indices_table.html").write_text(
        render_table(
            table_id="terminology-indices-table",
            headers=indices_headers,
            rows=indices_rows,
        ),
        encoding="utf-8",
    )

    (OUTPUT_DIR / "matrices_table.html").write_text(
        render_table(
            table_id="terminology-matrices-table",
            headers=matrices_headers,
            rows=matrices_rows,
        ),
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
